"""HOAi Customer Risk Monitor — Excel Generator.

Reads a daily-report JSON and generates a branded Excel workbook scoring
each customer on a composite 0-100 risk scale across 6 weighted signals:
deflection, CSAT, transfer rate, margin, volume trend, and error rate.

Usage:
    python daily-reports/generate-risk-monitor.py                        # Latest JSON
    python daily-reports/generate-risk-monitor.py --date 2026-04-27     # Specific date
"""

import argparse
import glob
import json
import os
import shutil
import sys
from datetime import datetime

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(SCRIPT_DIR, 'data')
OUTPUT_DIR = os.path.join(SCRIPT_DIR, 'output')
CONFIG_PATH = os.path.join(SCRIPT_DIR, 'daily-report-config.json')
ONEDRIVE_TARGET = os.path.join(
    os.path.expanduser('~'),
    'OneDrive - Vantaca, LLC', 'HOAi - Documents', 'Strategy & Ops',
    'Voice Usage Report', 'output',
)

os.makedirs(OUTPUT_DIR, exist_ok=True)

with open(CONFIG_PATH) as f:
    CONFIG = json.load(f)

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)

# ── Style constants (matching daily-report palette) ──────────────────────

NAVY = '1B2A4A'
BLUE = '2563EB'
GREEN = '059669'
AMBER = 'D97706'
RED = 'DC2626'
GRAY = '6B7280'
DARK = '111827'
WHITE = 'FFFFFF'
SURFACE = 'F3F4F6'

HEADER_FILL = PatternFill('solid', fgColor=NAVY)
HEADER_FONT = Font(name='Calibri', bold=True, color=WHITE, size=10)
BODY_FONT = Font(name='Calibri', color=DARK, size=10)
GOOD_FONT = Font(name='Calibri', color=GREEN, size=10, bold=True)
WARN_FONT = Font(name='Calibri', color=AMBER, size=10, bold=True)
BAD_FONT = Font(name='Calibri', color=RED, size=10, bold=True)
TOTAL_FILL = PatternFill('solid', fgColor=SURFACE)
TOTAL_FONT = Font(name='Calibri', bold=True, color=DARK, size=10)
MUTED_FONT = Font(name='Calibri', color=GRAY, size=9)
TITLE_FONT = Font(name='Calibri', bold=True, color=DARK, size=14)
SUBTITLE_FONT = Font(name='Calibri', color=GRAY, size=11)
THIN_BORDER = Border(bottom=Side(style='thin', color='E5E7EB'))
THICK_BORDER = Border(
    top=Side(style='medium', color=NAVY),
    bottom=Side(style='medium', color=NAVY),
)

TIER_STYLES = {
    'High':     (PatternFill('solid', fgColor='FEF2F2'), Font(name='Calibri', color=RED, size=10, bold=True)),
    'Elevated': (PatternFill('solid', fgColor='FFF7ED'), Font(name='Calibri', color=AMBER, size=10, bold=True)),
    'Watch':    (PatternFill('solid', fgColor='FFFBEB'), Font(name='Calibri', color='92400E', size=10, bold=True)),
    'Healthy':  (PatternFill('solid', fgColor='ECFDF5'), Font(name='Calibri', color=GREEN, size=10, bold=True)),
}

FMT_PCT = '0.0%'
FMT_PCT2 = '0.00%'
FMT_NUM = '#,##0'
FMT_DEC1 = '#,##0.0'
FMT_DEC2 = '0.00'
FMT_SIGN_PCT = '+0.0%;-0.0%'


# ── Helpers ──────────────────────────────────────────────────────────────

def _safe(val, default=0):
    return val if val is not None else default


def _clamp(val, lo=0, hi=None):
    if hi is None:
        return max(lo, val)
    return max(lo, min(hi, val))


def _apply_header(ws, row, col_count):
    for c in range(1, col_count + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)


def _apply_total_row(ws, row, col_count):
    for c in range(1, col_count + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = TOTAL_FILL
        cell.font = TOTAL_FONT
        cell.border = THICK_BORDER


def _auto_width(ws, min_width=10, max_width=32):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(min(max_len + 2, max_width), min_width)


def _rate_font(value, metric):
    benchmarks = CONFIG.get('voice_benchmarks', {})
    bench = benchmarks.get(metric, {})
    target = bench.get('target')
    warn = bench.get('warn')
    if target is None:
        return BODY_FONT
    if target > warn:  # higher is better (deflection, csat)
        if value >= target:
            return GOOD_FONT
        elif value >= warn:
            return WARN_FONT
        return BAD_FONT
    else:  # lower is better (transfer, error)
        if value <= target:
            return GOOD_FONT
        elif value <= warn:
            return WARN_FONT
        return BAD_FONT


# ── Risk scoring ─────────────────────────────────────────────────────────

BENCH = CONFIG.get('voice_benchmarks', {})

SIGNAL_WEIGHTS = {
    'Deflection':    30,
    'CSAT':          20,
    'Transfer Rate': 15,
    'Margin':        15,
    'Volume Trend':  10,
    'Error Rate':    10,
}

TIER_THRESHOLDS = [(50, 'High'), (40, 'Elevated'), (25, 'Watch'), (0, 'Healthy')]

ACTION_MAP = {
    ('High', 'Deflection'):    'Review AOP urgently - deflection critically low',
    ('High', 'CSAT'):          'Schedule QBR - CSAT below acceptable threshold',
    ('High', 'Transfer Rate'): 'Audit transfer config - excessive transfers',
    ('High', 'Margin'):        'Cost review needed - deeply negative margin',
    ('High', 'Volume Trend'):  'Engagement at risk - significant volume decline',
    ('High', 'Error Rate'):    'Escalate to engineering - high error rate',
    ('Elevated', 'Deflection'):    'AOP tune-up needed - deflection below target',
    ('Elevated', 'CSAT'):          'Review worst-CSAT calls - satisfaction declining',
    ('Elevated', 'Transfer Rate'): 'Check transfer destinations - rate above warn',
    ('Elevated', 'Margin'):        'Review COGS profile - margin under pressure',
    ('Elevated', 'Volume Trend'):  'Check in with customer - volume dropping',
    ('Elevated', 'Error Rate'):    'Investigate error patterns - rate elevated',
    ('Watch', 'Deflection'):   'Monitor deflection trend next 2 weeks',
    ('Watch', 'CSAT'):         'Monitor CSAT - approaching warn threshold',
    ('Watch', 'Transfer Rate'):'Monitor transfer rate trend',
    ('Watch', 'Margin'):       'Monitor margin trend',
    ('Watch', 'Volume Trend'): 'Monitor volume - slight decline noted',
    ('Watch', 'Error Rate'):   'Monitor error rate trend',
}


def compute_risk(company_data):
    """Compute composite risk score (0-100) and return (score, tier, top_signal, action)."""
    defl = _safe(company_data.get('deflection_rate'), 0)
    csat = _safe(company_data.get('avg_csat'), 4.5)
    xfer = _safe(company_data.get('transfer_rate'), 0)
    margin = _safe(company_data.get('margin_pct'), 0)
    vs_prior = _safe(company_data.get('vs_prior_month'), 0)
    error = _safe(company_data.get('error_rate_actionable'), 0)

    defl_target = BENCH.get('deflection_rate', {}).get('target', 0.60)
    csat_target = BENCH.get('csat', {}).get('target', 4.5)
    xfer_warn = BENCH.get('transfer_rate', {}).get('warn', 0.40)
    margin_target = BENCH.get('margin_pct', {}).get('target', 0.60)
    margin_warn = BENCH.get('margin_pct', {}).get('warn', 0.40)
    error_warn = BENCH.get('error_rate_actionable', {}).get('warn', 0.08)

    scores = {}

    # Deflection: 30 pts — lower deflection = higher risk
    scores['Deflection'] = _clamp(30 * (1 - defl / defl_target), 0, 30) if defl_target > 0 else 0

    # CSAT: 20 pts — lower CSAT = higher risk
    scores['CSAT'] = _clamp(20 * (1 - csat / csat_target), 0, 20) if csat_target > 0 else 0

    # Transfer: 15 pts — higher transfer = higher risk
    scores['Transfer Rate'] = _clamp(15 * (xfer / xfer_warn), 0, 15) if xfer_warn > 0 else 0

    # Margin: 15 pts — step function
    if margin < margin_warn:
        scores['Margin'] = 15
    elif margin < margin_target:
        scores['Margin'] = 7
    else:
        scores['Margin'] = 0

    # Volume trend: 10 pts — only penalize declining
    if vs_prior < 0:
        scores['Volume Trend'] = _clamp(10 * abs(vs_prior), 0, 10)
    else:
        scores['Volume Trend'] = 0

    # Error rate: 10 pts — higher error = higher risk
    scores['Error Rate'] = _clamp(10 * (error / error_warn), 0, 10) if error_warn > 0 else 0

    total = sum(scores.values())
    top_signal = max(scores, key=scores.get)

    tier = 'Healthy'
    for threshold, label in TIER_THRESHOLDS:
        if total >= threshold:
            tier = label
            break

    action = ACTION_MAP.get((tier, top_signal), '')

    return round(total, 1), tier, top_signal, action


# ── Excel generation ─────────────────────────────────────────────────────

HEADERS = [
    'Company', 'Risk Tier', 'Risk Score', 'Top Risk Signal',
    'Calls MTD', 'Daily Avg', 'vs Prior Mo',
    'Deflection %', 'Transfer %', 'Avg CSAT', 'Error Rate',
    'Margin %', 'Utilization', 'Included Calls', 'Projected EOM',
    'Hours Saved', 'Action',
]
COL_COUNT = len(HEADERS)


def generate_risk_monitor(data, output_path):
    """Generate the risk monitor Excel workbook."""
    report_date = data.get('report_date', 'unknown')
    voice = data.get('voice', {})
    platform = data.get('platform', {})
    vs = platform.get('voice_summary', {})

    # Build rows
    rows = []
    for name, v in voice.items():
        if not isinstance(v, dict) or 'total_calls' not in v:
            continue
        score, tier, top_signal, action = compute_risk(v)
        ri = v.get('revenue_intel') or {}
        rows.append({
            'company': name,
            'tier': tier,
            'score': score,
            'top_signal': top_signal,
            'calls': _safe(v.get('total_calls')),
            'daily_avg': _safe(v.get('daily_avg')),
            'vs_prior': _safe(v.get('vs_prior_month')),
            'deflection': _safe(v.get('deflection_rate')),
            'transfer': _safe(v.get('transfer_rate')),
            'csat': _safe(v.get('avg_csat')),
            'error': _safe(v.get('error_rate_actionable')),
            'margin': _safe(v.get('margin_pct')),
            'utilization': ri.get('flag', ''),
            'included': ri.get('included') or '',
            'projected': ri.get('projected_eom') or '',
            'hours_saved': _safe(v.get('hours_saved')),
            'action': action,
        })

    rows.sort(key=lambda r: r['score'], reverse=True)

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = 'Risk Monitor'

    # Title block
    ws.cell(row=1, column=1, value='HOAi Customer Risk Monitor').font = TITLE_FONT
    days = vs.get('days_elapsed', '')
    dt = datetime.strptime(report_date, '%Y-%m-%d') if report_date != 'unknown' else datetime.now()
    month_name = dt.strftime('%B')
    ws.cell(row=2, column=1,
            value=f'Report Date: {report_date}  |  Period: {month_name} 1 - {report_date} ({days} days MTD)').font = SUBTITLE_FONT

    # Headers at row 4
    hdr_row = 4
    for i, h in enumerate(HEADERS, 1):
        ws.cell(row=hdr_row, column=i, value=h)
    _apply_header(ws, hdr_row, COL_COUNT)

    # Data rows
    tier_counts = {'High': 0, 'Elevated': 0, 'Watch': 0, 'Healthy': 0}
    for idx, r in enumerate(rows):
        row_num = hdr_row + 1 + idx
        tier_counts[r['tier']] = tier_counts.get(r['tier'], 0) + 1

        tier_fill, tier_font = TIER_STYLES.get(r['tier'], (None, BODY_FONT))

        # A: Company
        c = ws.cell(row=row_num, column=1, value=r['company'])
        c.font = BODY_FONT
        c.border = THIN_BORDER

        # B: Risk Tier
        c = ws.cell(row=row_num, column=2, value=r['tier'])
        c.font = tier_font
        if tier_fill:
            c.fill = tier_fill
        c.alignment = Alignment(horizontal='center')
        c.border = THIN_BORDER

        # C: Risk Score
        c = ws.cell(row=row_num, column=3, value=r['score'])
        c.font = tier_font
        c.alignment = Alignment(horizontal='center')
        c.border = THIN_BORDER

        # D: Top Risk Signal
        c = ws.cell(row=row_num, column=4, value=r['top_signal'])
        c.font = BODY_FONT
        c.border = THIN_BORDER

        # E: Calls MTD
        c = ws.cell(row=row_num, column=5, value=r['calls'])
        c.font = BODY_FONT
        c.number_format = FMT_NUM
        c.alignment = Alignment(horizontal='right')
        c.border = THIN_BORDER

        # F: Daily Avg
        c = ws.cell(row=row_num, column=6, value=r['daily_avg'])
        c.font = BODY_FONT
        c.number_format = FMT_DEC1
        c.alignment = Alignment(horizontal='right')
        c.border = THIN_BORDER

        # G: vs Prior Mo
        c = ws.cell(row=row_num, column=7, value=r['vs_prior'] / 100 if r['vs_prior'] else 0)
        c.font = GOOD_FONT if r['vs_prior'] >= 0 else BAD_FONT
        c.number_format = FMT_SIGN_PCT
        c.alignment = Alignment(horizontal='right')
        c.border = THIN_BORDER

        # H: Deflection %
        c = ws.cell(row=row_num, column=8, value=r['deflection'])
        c.font = _rate_font(r['deflection'], 'deflection_rate')
        c.number_format = FMT_PCT
        c.alignment = Alignment(horizontal='right')
        c.border = THIN_BORDER

        # I: Transfer %
        c = ws.cell(row=row_num, column=9, value=r['transfer'])
        c.font = _rate_font(r['transfer'], 'transfer_rate')
        c.number_format = FMT_PCT
        c.alignment = Alignment(horizontal='right')
        c.border = THIN_BORDER

        # J: Avg CSAT
        c = ws.cell(row=row_num, column=10, value=r['csat'])
        c.font = _rate_font(r['csat'], 'csat')
        c.number_format = FMT_DEC2
        c.alignment = Alignment(horizontal='right')
        c.border = THIN_BORDER

        # K: Error Rate
        c = ws.cell(row=row_num, column=11, value=r['error'])
        c.font = _rate_font(r['error'], 'error_rate_actionable')
        c.number_format = FMT_PCT2
        c.alignment = Alignment(horizontal='right')
        c.border = THIN_BORDER

        # L: Margin %
        c = ws.cell(row=row_num, column=12, value=r['margin'] / 100 if r['margin'] else 0)
        c.font = GOOD_FONT if _safe(r['margin']) >= 0 else BAD_FONT
        c.number_format = FMT_SIGN_PCT
        c.alignment = Alignment(horizontal='right')
        c.border = THIN_BORDER

        # M: Utilization
        c = ws.cell(row=row_num, column=13, value=r['utilization'])
        c.font = BODY_FONT
        c.alignment = Alignment(horizontal='center')
        c.border = THIN_BORDER

        # N: Included Calls
        c = ws.cell(row=row_num, column=14, value=r['included'])
        c.font = BODY_FONT
        c.number_format = FMT_NUM
        c.alignment = Alignment(horizontal='right')
        c.border = THIN_BORDER

        # O: Projected EOM
        c = ws.cell(row=row_num, column=15, value=r['projected'])
        c.font = BODY_FONT
        c.number_format = FMT_NUM
        c.alignment = Alignment(horizontal='right')
        c.border = THIN_BORDER

        # P: Hours Saved
        c = ws.cell(row=row_num, column=16, value=r['hours_saved'])
        c.font = BODY_FONT
        c.number_format = FMT_DEC1
        c.alignment = Alignment(horizontal='right')
        c.border = THIN_BORDER

        # Q: Action
        c = ws.cell(row=row_num, column=17, value=r['action'])
        c.font = MUTED_FONT
        c.border = THIN_BORDER

    # Totals row
    total_row = hdr_row + 1 + len(rows)
    ws.cell(row=total_row, column=1, value='PLATFORM TOTALS')
    ws.cell(row=total_row, column=5, value=vs.get('total_calls', 0))
    ws.cell(row=total_row, column=5).number_format = FMT_NUM
    avg_daily = round(vs.get('total_calls', 0) / max(vs.get('days_elapsed', 1), 1), 1)
    ws.cell(row=total_row, column=6, value=avg_daily)
    ws.cell(row=total_row, column=6).number_format = FMT_DEC1
    ws.cell(row=total_row, column=8, value=vs.get('deflection_rate', 0))
    ws.cell(row=total_row, column=8).number_format = FMT_PCT
    ws.cell(row=total_row, column=9, value=vs.get('transfer_rate', 0))
    ws.cell(row=total_row, column=9).number_format = FMT_PCT
    ws.cell(row=total_row, column=10, value=vs.get('avg_csat', 0))
    ws.cell(row=total_row, column=10).number_format = FMT_DEC2
    ws.cell(row=total_row, column=16, value=vs.get('hours_saved', 0))
    ws.cell(row=total_row, column=16).number_format = FMT_DEC1
    _apply_total_row(ws, total_row, COL_COUNT)

    # Benchmark reference row
    bench_row = total_row + 1
    ws.cell(row=bench_row, column=1, value='BENCHMARKS (target)').font = MUTED_FONT
    ws.cell(row=bench_row, column=8, value=BENCH.get('deflection_rate', {}).get('target', '')).font = MUTED_FONT
    ws.cell(row=bench_row, column=8).number_format = FMT_PCT
    ws.cell(row=bench_row, column=9, value=BENCH.get('transfer_rate', {}).get('target', '')).font = MUTED_FONT
    ws.cell(row=bench_row, column=9).number_format = FMT_PCT
    ws.cell(row=bench_row, column=10, value=BENCH.get('csat', {}).get('target', '')).font = MUTED_FONT
    ws.cell(row=bench_row, column=10).number_format = FMT_DEC2
    ws.cell(row=bench_row, column=11, value=BENCH.get('error_rate_actionable', {}).get('target', '')).font = MUTED_FONT
    ws.cell(row=bench_row, column=11).number_format = FMT_PCT2

    # Tier summary row
    summary_row = bench_row + 2
    ws.cell(row=summary_row, column=1, value='Tier Summary:').font = Font(name='Calibri', bold=True, color=DARK, size=10)
    col = 2
    for tier_name in ['High', 'Elevated', 'Watch', 'Healthy']:
        count = tier_counts.get(tier_name, 0)
        c = ws.cell(row=summary_row, column=col, value=f'{tier_name}: {count}')
        _, tfont = TIER_STYLES.get(tier_name, (None, BODY_FONT))
        c.font = tfont
        col += 1

    # Freeze panes and auto-width
    ws.freeze_panes = f'A{hdr_row + 1}'
    _auto_width(ws)
    # Widen action column
    ws.column_dimensions[get_column_letter(17)].width = 48

    wb.save(output_path)
    return tier_counts


# ── CLI ──────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description='Generate HOAi Customer Risk Monitor Excel')
    parser.add_argument('--date', help='Report date (YYYY-MM-DD). Default: latest JSON in data/')
    parser.add_argument('--json', help='Path to daily-report JSON file')
    args = parser.parse_args()

    # Resolve JSON path
    if args.json:
        json_path = args.json
    elif args.date:
        json_path = os.path.join(DATA_DIR, f'daily-report-{args.date}.json')
    else:
        files = sorted(glob.glob(os.path.join(DATA_DIR, 'daily-report-*.json')))
        if not files:
            print('ERROR: No daily-report JSON files found in data/')
            sys.exit(1)
        json_path = files[-1]

    if not os.path.exists(json_path):
        print(f'ERROR: JSON not found: {json_path}')
        sys.exit(1)

    with open(json_path) as f:
        data = json.load(f)

    report_date = data.get('report_date', 'unknown')
    output_path = os.path.join(OUTPUT_DIR, f'HOAi_Risk_Monitor_{report_date}.xlsx')

    print(f'[Risk Monitor] Generating for {report_date}...')
    tier_counts = generate_risk_monitor(data, output_path)
    print(f'[Risk Monitor] Saved: {output_path}')
    print(f'[Risk Monitor] High: {tier_counts.get("High", 0)}  '
          f'Elevated: {tier_counts.get("Elevated", 0)}  '
          f'Watch: {tier_counts.get("Watch", 0)}  '
          f'Healthy: {tier_counts.get("Healthy", 0)}')

    # Copy to OneDrive
    try:
        os.makedirs(ONEDRIVE_TARGET, exist_ok=True)
        dest = os.path.join(ONEDRIVE_TARGET, os.path.basename(output_path))
        shutil.copy2(output_path, dest)
        print(f'[Risk Monitor] Copied to OneDrive: {dest}')
    except Exception as e:
        print(f'[Risk Monitor] Warning: OneDrive copy failed: {e}')


if __name__ == '__main__':
    main()
