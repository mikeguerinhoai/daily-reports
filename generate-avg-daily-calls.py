"""Generate detailed Excel: April MTD Average Daily Call Volume per Customer."""
import json, os, sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference

DATE = "2026-04-28"
DATA_PATH = f"daily-reports/data/daily-report-{DATE}.json"
OUT_DIR = f"daily-reports/output/2026-04/{DATE}/Dashboard"
OUT_FILE = f"{OUT_DIR}/HOAi_April_MTD_Avg_Daily_Calls_{DATE}.xlsx"

d = json.load(open(DATA_PATH))
days = d["platform"]["voice_summary"].get("days_elapsed", 28)
voice = d["voice"]
rev_intel = {
    r["company"]: r
    for r in d.get("revenue_intelligence", [])
    if r.get("channel") == "voice"
}

# ── Build rows ──
rows = []
for name, data in voice.items():
    if not isinstance(data, dict):
        continue
    calls = data.get("total_calls", 0)
    deflection = data.get("deflection_rate", data.get("deflection", 0))
    transfer = data.get("transfer_rate", data.get("transfer", 0))
    avg_duration = data.get("avg_duration", data.get("avg_call_duration", 0))
    hours_saved = data.get("hours_saved", 0)

    ri = rev_intel.get(name, {})
    included = ri.get("included", 0)
    pace_pct = ri.get("pace_pct", 0)
    projected = ri.get("projected_eom", 0)

    avg_day = round(calls / days, 1) if days > 0 else 0

    if avg_day >= 100:
        bucket = "Tier 1 (100+/day)"
    elif avg_day >= 25:
        bucket = "Tier 2 (25-99/day)"
    elif avg_day >= 5:
        bucket = "Tier 3 (5-24/day)"
    else:
        bucket = "Tier 4 (<5/day)"

    rows.append(dict(
        name=name, calls=calls, avg_day=avg_day,
        deflection=deflection, transfer=transfer,
        avg_duration=avg_duration or 0, hours_saved=hours_saved,
        included=included, pace_pct=pace_pct,
        projected=projected, bucket=bucket,
    ))

rows.sort(key=lambda x: -x["calls"])
total_calls = sum(r["calls"] for r in rows)

# ── Styles ──
NAVY = "1B2A4A"
TEAL = "0891B2"
LIGHT_GRAY = "F8FAFC"

header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
subheader_fill = PatternFill(start_color=TEAL, end_color=TEAL, fill_type="solid")
alt_fill = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid")
totals_fill = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")
green_fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
yellow_fill = PatternFill(start_color="FEF9C3", end_color="FEF9C3", fill_type="solid")
orange_fill = PatternFill(start_color="FFEDD5", end_color="FFEDD5", fill_type="solid")
red_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
tier_fills = {
    "Tier 1 (100+/day)": green_fill,
    "Tier 2 (25-99/day)": yellow_fill,
    "Tier 3 (5-24/day)": orange_fill,
    "Tier 4 (<5/day)": red_fill,
}
thin_border = Border(
    left=Side(style="thin", color="E2E8F0"),
    right=Side(style="thin", color="E2E8F0"),
    top=Side(style="thin", color="E2E8F0"),
    bottom=Side(style="thin", color="E2E8F0"),
)

def style_header_row(ws, row, ncols, fill=None):
    for col in range(1, ncols + 1):
        c = ws.cell(row=row, column=col)
        c.font = header_font
        c.fill = fill or header_fill
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = thin_border

def base_style(cell, align="center", fmt=None):
    cell.border = thin_border
    cell.alignment = Alignment(horizontal=align, vertical="center")
    if fmt:
        cell.number_format = fmt
    return cell

def pct_val(raw):
    """Normalize: if stored as decimal <1 convert to %, else keep."""
    if raw and raw < 1:
        return round(raw * 100, 1)
    return round(raw, 1) if raw else 0

# ═════════════════════════════════════════════════
wb = Workbook()

# ── TAB 1: Customer Detail ──
ws = wb.active
ws.title = "Customer Detail"
ws.sheet_properties.tabColor = NAVY

ws.merge_cells("A1:L1")
c = ws["A1"]
c.value = f"HOAi Voice — April MTD Avg Daily Call Volume (Apr 1\u201328, {days} days)"
c.font = Font(name="Calibri", bold=True, size=14, color=NAVY)
c.alignment = Alignment(horizontal="left", vertical="center")
ws.row_dimensions[1].height = 30

ws.merge_cells("A2:L2")
ws["A2"].value = (
    f"Generated 2026-04-29  |  {len(rows)} active companies  |  "
    f"{total_calls:,} total calls  |  {round(total_calls/days,1):,} avg/day"
)
ws["A2"].font = Font(name="Calibri", size=10, color="64748B")

headers = [
    "#", "Company", "Volume Tier", "MTD Calls", "Avg Calls/Day",
    "Deflection %", "Transfer %", "Avg Duration (s)", "Hours Saved",
    "Included Calls", "Pace vs Plan %", "Projected EOM",
]
for col, h in enumerate(headers, 1):
    ws.cell(row=4, column=col, value=h)
style_header_row(ws, 4, len(headers))
ws.row_dimensions[4].height = 32

col_fmts = {
    1: (None, "center"), 2: (None, "left"), 3: (None, "center"),
    4: ("#,##0", "center"), 5: ("#,##0.0", "center"),
    6: ("0.0", "center"), 7: ("0.0", "center"),
    8: ("0.0", "center"), 9: ("#,##0.0", "center"),
    10: ("#,##0", "center"), 11: ("0.0", "center"), 12: ("#,##0", "center"),
}

for i, r in enumerate(rows):
    row = i + 5
    vals = [
        i + 1, r["name"], r["bucket"], r["calls"], r["avg_day"],
        pct_val(r["deflection"]), pct_val(r["transfer"]),
        round(r["avg_duration"], 1), round(r["hours_saved"], 1),
        r["included"] or "", pct_val(r["pace_pct"]) or "",
        round(r["projected"]) if r["projected"] else "",
    ]
    for col, v in enumerate(vals, 1):
        cell = ws.cell(row=row, column=col, value=v)
        fmt, align = col_fmts.get(col, (None, "center"))
        base_style(cell, align=align, fmt=fmt)

    if i % 2 == 1:
        for col in range(1, len(headers) + 1):
            ws.cell(row=row, column=col).fill = alt_fill

    ws.cell(row=row, column=3).fill = tier_fills.get(r["bucket"], alt_fill)

# Totals
tr = len(rows) + 5
for col in range(1, len(headers) + 1):
    c = ws.cell(row=tr, column=col)
    c.font = Font(name="Calibri", bold=True, size=11)
    c.fill = totals_fill
    c.border = thin_border
    c.alignment = Alignment(horizontal="center", vertical="center")
ws.cell(row=tr, column=2, value="TOTAL").alignment = Alignment(horizontal="left", vertical="center")
base_style(ws.cell(row=tr, column=4, value=total_calls), fmt="#,##0")
base_style(ws.cell(row=tr, column=5, value=round(total_calls / days, 1)), fmt="#,##0.0")
base_style(ws.cell(row=tr, column=9, value=round(sum(r["hours_saved"] for r in rows), 1)), fmt="#,##0.0")
ws.cell(row=tr, column=4).font = Font(name="Calibri", bold=True, size=11)
ws.cell(row=tr, column=5).font = Font(name="Calibri", bold=True, size=11)

widths = [5, 42, 20, 12, 15, 14, 12, 17, 12, 15, 15, 14]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = "C5"
ws.auto_filter.ref = f"A4:L{tr - 1}"

# ── TAB 2: Tier Summary ──
ws2 = wb.create_sheet("Tier Summary")
ws2.sheet_properties.tabColor = TEAL

tier_order = ["Tier 1 (100+/day)", "Tier 2 (25-99/day)", "Tier 3 (5-24/day)", "Tier 4 (<5/day)"]
tiers = {}
for r in rows:
    b = r["bucket"]
    tiers.setdefault(b, {"count": 0, "calls": 0, "hours": 0, "companies": []})
    tiers[b]["count"] += 1
    tiers[b]["calls"] += r["calls"]
    tiers[b]["hours"] += r["hours_saved"]
    tiers[b]["companies"].append(r["name"])

ws2.merge_cells("A1:G1")
ws2["A1"].value = "Volume Tier Summary — April MTD"
ws2["A1"].font = Font(name="Calibri", bold=True, size=14, color=NAVY)
ws2.row_dimensions[1].height = 30

t_headers = ["Tier", "Companies", "% of Book", "MTD Calls", "% of Calls", "Avg/Day (Tier)", "Hours Saved"]
for col, h in enumerate(t_headers, 1):
    ws2.cell(row=3, column=col, value=h)
style_header_row(ws2, 3, len(t_headers))

tier_fill_list = [green_fill, yellow_fill, orange_fill, red_fill]
for i, tier in enumerate(tier_order):
    row = i + 4
    td = tiers.get(tier, {"count": 0, "calls": 0, "hours": 0})
    base_style(ws2.cell(row=row, column=1, value=tier), align="left")
    ws2.cell(row=row, column=1).fill = tier_fill_list[i]
    base_style(ws2.cell(row=row, column=2, value=td["count"]))
    base_style(ws2.cell(row=row, column=3, value=round(td["count"] / len(rows) * 100, 1)), fmt="0.0")
    base_style(ws2.cell(row=row, column=4, value=td["calls"]), fmt="#,##0")
    base_style(ws2.cell(row=row, column=5, value=round(td["calls"] / total_calls * 100, 1) if total_calls else 0), fmt="0.0")
    base_style(ws2.cell(row=row, column=6, value=round(td["calls"] / days, 1)), fmt="#,##0.0")
    base_style(ws2.cell(row=row, column=7, value=round(td["hours"], 1)), fmt="#,##0.0")

# Company lists per tier
r_off = 10
ws2.cell(row=r_off, column=1, value="Tier Membership").font = Font(name="Calibri", bold=True, size=12, color=NAVY)
for i, tier in enumerate(tier_order):
    r = r_off + 1 + i
    ws2.cell(row=r, column=1, value=tier).font = Font(name="Calibri", bold=True, size=10)
    ws2.cell(row=r, column=1).fill = tier_fill_list[i]
    ws2.merge_cells(start_row=r, start_column=2, end_row=r, end_column=7)
    c = ws2.cell(row=r, column=2, value=", ".join(tiers.get(tier, {}).get("companies", [])))
    c.alignment = Alignment(wrap_text=True, vertical="top")
    ws2.row_dimensions[r].height = 40

for col, w in enumerate([24, 14, 12, 14, 12, 16, 14], 1):
    ws2.column_dimensions[get_column_letter(col)].width = w

# ── TAB 3: Top 20 Chart ──
ws3 = wb.create_sheet("Top 20 Chart")
ws3.sheet_properties.tabColor = "059669"

ws3.merge_cells("A1:C1")
ws3["A1"].value = "Top 20 Companies by Avg Daily Calls — April MTD"
ws3["A1"].font = Font(name="Calibri", bold=True, size=14, color=NAVY)

ch_headers = ["Company", "Avg Calls/Day", "MTD Calls"]
for col, h in enumerate(ch_headers, 1):
    ws3.cell(row=3, column=col, value=h)
style_header_row(ws3, 3, 3)

for i, r in enumerate(rows[:20]):
    row = i + 4
    base_style(ws3.cell(row=row, column=1, value=r["name"]), align="left")
    base_style(ws3.cell(row=row, column=2, value=r["avg_day"]), fmt="#,##0.0")
    base_style(ws3.cell(row=row, column=3, value=r["calls"]), fmt="#,##0")

chart = BarChart()
chart.type = "bar"
chart.style = 10
chart.title = "Top 20 — Avg Calls per Day (April MTD)"
chart.y_axis.title = "Avg Calls/Day"
chart.width = 28
chart.height = 16
data_ref = Reference(ws3, min_col=2, min_row=3, max_row=23)
cats_ref = Reference(ws3, min_col=1, min_row=4, max_row=23)
chart.add_data(data_ref, titles_from_data=True)
chart.set_categories(cats_ref)
chart.series[0].graphicalProperties.solidFill = "0891B2"
ws3.add_chart(chart, "E3")

ws3.column_dimensions["A"].width = 42
ws3.column_dimensions["B"].width = 16
ws3.column_dimensions["C"].width = 12

# ── Save ──
os.makedirs(OUT_DIR, exist_ok=True)
wb.save(OUT_FILE)
print(f"Saved: {OUT_FILE}")
print(f"Tabs:  Customer Detail | Tier Summary | Top 20 Chart")
print(f"Rows:  {len(rows)} companies, {total_calls:,} calls, {days} days MTD")
