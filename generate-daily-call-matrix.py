"""Generate April MTD Daily Call Volume Matrix per Customer (Excel).

Tabs:
  1. Daily Call Matrix  — company x date grid with heat-map
  2. Activation Status  — active days, streaks, status flags
  3. Top 20 Chart       — horizontal bar by avg/active day
"""

import os, sys
from datetime import date, timedelta
from collections import defaultdict

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import ColorScaleRule

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from supabase.queries.call_logs import get_book_daily_volume
from supabase.queries.management_companies import list_active

# ── Config ──
START = date(2026, 4, 1)
END = date(2026, 4, 28)
TOTAL_DAYS = (END - START).days + 1  # 28
DATES = [START + timedelta(days=i) for i in range(TOTAL_DAYS)]

OUT_DIR = f"daily-reports/output/2026-04/{END}/Dashboard"
OUT_FILE = f"{OUT_DIR}/HOAi_April_Daily_Call_Matrix_{END}.xlsx"

# ── Fetch ──
print(f"Fetching daily volume: {START} to {END} ...")
raw = get_book_daily_volume(str(START), str(END), channel="voice")
print(f"  Call rows: {len(raw)}")

print("Fetching active management companies ...")
all_active = list_active()
all_names = sorted(set(c["name"] for c in all_active))
print(f"  Active companies: {len(all_names)}")

# ── Pivot: company -> {date: total} ──
grid = defaultdict(lambda: defaultdict(int))
for r in raw:
    d = str(r["date"])[:10]
    grid[r["company_name"]][d] += r["total"]

# Ensure every active company is in the grid (with empty day_map if no calls)
for name in all_names:
    if name not in grid:
        grid[name]  # creates empty defaultdict(int)

# ── Compute per-company stats ──
companies = []
for name, day_map in grid.items():
    mtd = sum(day_map.values())
    active_days = sum(1 for dt in DATES if day_map.get(str(dt), 0) > 0)
    avg_active = round(mtd / active_days, 1) if active_days else 0
    avg_calendar = round(mtd / TOTAL_DAYS, 1)

    # Longest consecutive streak
    streak = max_streak = 0
    for dt in DATES:
        if day_map.get(str(dt), 0) > 0:
            streak += 1
            max_streak = max(max_streak, streak)
        else:
            streak = 0

    # First call date
    first_call = None
    for dt in DATES:
        if day_map.get(str(dt), 0) > 0:
            first_call = dt
            break

    # Status
    active_pct = active_days / TOTAL_DAYS if TOTAL_DAYS else 0
    if active_days == 0:
        status = "Inactive"
    elif active_pct >= 0.9:
        status = "Full Month"
    elif active_pct >= 0.5:
        status = "Partial"
    elif first_call and first_call >= date(2026, 4, 8):
        status = "New Activation"
    else:
        status = "Sporadic"

    companies.append(dict(
        name=name, mtd=mtd, active_days=active_days,
        active_pct=round(active_pct * 100, 1),
        avg_active=avg_active, avg_calendar=avg_calendar,
        max_streak=max_streak, first_call=first_call,
        status=status, day_map=day_map,
    ))

companies.sort(key=lambda c: -c["mtd"])
total_calls = sum(c["mtd"] for c in companies)
print(f"  Companies: {len(companies)}, Total calls: {total_calls:,}")

# ═══════════════════════════════════════════════
# Styles
# ═══════════════════════════════════════════════
NAVY = "1B2A4A"
TEAL = "0891B2"
LIGHT_GRAY = "F8FAFC"

hdr_font = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
hdr_fill = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
alt_fill = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid")
totals_fill = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")
green_fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
yellow_fill = PatternFill(start_color="FEF9C3", end_color="FEF9C3", fill_type="solid")
orange_fill = PatternFill(start_color="FFEDD5", end_color="FFEDD5", fill_type="solid")
red_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
gray_fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
status_fills = {
    "Full Month": green_fill,
    "Partial": yellow_fill,
    "New Activation": orange_fill,
    "Sporadic": red_fill,
    "Inactive": gray_fill,
}
ALL_STATUSES = ["Full Month", "Partial", "New Activation", "Sporadic", "Inactive"]
thin = Border(
    left=Side(style="thin", color="E2E8F0"),
    right=Side(style="thin", color="E2E8F0"),
    top=Side(style="thin", color="E2E8F0"),
    bottom=Side(style="thin", color="E2E8F0"),
)

def sc(cell, align="center", fmt=None):
    cell.border = thin
    cell.alignment = Alignment(horizontal=align, vertical="center")
    if fmt:
        cell.number_format = fmt
    return cell

def header_row(ws, row, ncols, fill=None):
    for col in range(1, ncols + 1):
        c = ws.cell(row=row, column=col)
        c.font = hdr_font
        c.fill = fill or hdr_fill
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = thin

wb = Workbook()

# ═══════════════════════════════════════════════
# TAB 1: Daily Call Matrix
# ═══════════════════════════════════════════════
ws = wb.active
ws.title = "Daily Call Matrix"
ws.sheet_properties.tabColor = NAVY

# Title
ncols = 1 + TOTAL_DAYS + 3  # company + 28 days + MTD + Active Days + Avg/Active
ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
t = ws["A1"]
t.value = f"HOAi Voice \u2014 Daily Call Volume Matrix (April 1\u201328, 2026)"
t.font = Font(name="Calibri", bold=True, size=13, color=NAVY)
t.alignment = Alignment(horizontal="left", vertical="center")
ws.row_dimensions[1].height = 28

ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
ws["A2"].value = f"{len(companies)} companies  |  {total_calls:,} total calls  |  {TOTAL_DAYS} calendar days"
ws["A2"].font = Font(name="Calibri", size=9, color="64748B")

# Headers (row 4)
HR = 4
ws.cell(row=HR, column=1, value="Company")
for i, dt in enumerate(DATES):
    # Show as "4/1", "4/2", etc.
    ws.cell(row=HR, column=2 + i, value=dt.strftime("%-m/%d") if os.name != "nt" else dt.strftime("%#m/%d"))
ws.cell(row=HR, column=2 + TOTAL_DAYS, value="MTD Total")
ws.cell(row=HR, column=3 + TOTAL_DAYS, value="Active Days")
ws.cell(row=HR, column=4 + TOTAL_DAYS, value="Avg/Active Day")
header_row(ws, HR, ncols)
ws.row_dimensions[HR].height = 28

# Data rows
for i, c in enumerate(companies):
    row = HR + 1 + i
    sc(ws.cell(row=row, column=1, value=c["name"]), align="left")
    for j, dt in enumerate(DATES):
        val = c["day_map"].get(str(dt), 0)
        cell = sc(ws.cell(row=row, column=2 + j, value=val), fmt="#,##0")
        cell.font = Font(name="Calibri", size=9)
    sc(ws.cell(row=row, column=2 + TOTAL_DAYS, value=c["mtd"]), fmt="#,##0")
    ws.cell(row=row, column=2 + TOTAL_DAYS).font = Font(name="Calibri", bold=True, size=10)
    sc(ws.cell(row=row, column=3 + TOTAL_DAYS, value=c["active_days"]))
    sc(ws.cell(row=row, column=4 + TOTAL_DAYS, value=c["avg_active"]), fmt="#,##0.0")

    # Alternating rows
    if i % 2 == 1:
        for col in range(1, ncols + 1):
            ws.cell(row=row, column=col).fill = alt_fill

# Totals row
tr = HR + 1 + len(companies)
ws.cell(row=tr, column=1, value="TOTAL").font = Font(name="Calibri", bold=True)
for j, dt in enumerate(DATES):
    day_total = sum(c["day_map"].get(str(dt), 0) for c in companies)
    cell = sc(ws.cell(row=tr, column=2 + j, value=day_total), fmt="#,##0")
    cell.font = Font(name="Calibri", bold=True, size=9)
sc(ws.cell(row=tr, column=2 + TOTAL_DAYS, value=total_calls), fmt="#,##0")
ws.cell(row=tr, column=2 + TOTAL_DAYS).font = Font(name="Calibri", bold=True)
for col in range(1, ncols + 1):
    ws.cell(row=tr, column=col).fill = totals_fill
    ws.cell(row=tr, column=col).border = thin

# Heat-map color scale on daily cells (white -> light teal -> dark teal)
if len(companies) > 0:
    first_data_row = HR + 1
    last_data_row = HR + len(companies)
    first_day_col = get_column_letter(2)
    last_day_col = get_column_letter(1 + TOTAL_DAYS)
    rng = f"{first_day_col}{first_data_row}:{last_day_col}{last_data_row}"
    ws.conditional_formatting.add(rng, ColorScaleRule(
        start_type="num", start_value=0, start_color="FFFFFF",
        mid_type="percentile", mid_value=50, mid_color="B2F5EA",
        end_type="max", end_color="0891B2",
    ))

# Column widths
ws.column_dimensions["A"].width = 40
for j in range(TOTAL_DAYS):
    ws.column_dimensions[get_column_letter(2 + j)].width = 7
ws.column_dimensions[get_column_letter(2 + TOTAL_DAYS)].width = 12
ws.column_dimensions[get_column_letter(3 + TOTAL_DAYS)].width = 12
ws.column_dimensions[get_column_letter(4 + TOTAL_DAYS)].width = 14

ws.freeze_panes = "B5"
ws.auto_filter.ref = f"A{HR}:{get_column_letter(ncols)}{tr - 1}"

# ═══════════════════════════════════════════════
# TAB 2: Activation Status
# ═══════════════════════════════════════════════
ws2 = wb.create_sheet("Activation Status")
ws2.sheet_properties.tabColor = TEAL

ws2.merge_cells("A1:I1")
ws2["A1"].value = "Activation Status \u2014 April 2026"
ws2["A1"].font = Font(name="Calibri", bold=True, size=13, color=NAVY)
ws2.row_dimensions[1].height = 28

hdrs = [
    "#", "Company", "MTD Calls", "Active Days", f"Total Days ({TOTAL_DAYS})",
    "Active %", "Avg/Active Day", "Avg/Calendar Day", "Longest Streak", "First Call", "Status",
]
for col, h in enumerate(hdrs, 1):
    ws2.cell(row=3, column=col, value=h)
header_row(ws2, 3, len(hdrs))
ws2.row_dimensions[3].height = 28

fmts = {
    1: (None, "center"), 2: (None, "left"), 3: ("#,##0", "center"),
    4: (None, "center"), 5: (None, "center"), 6: ("0.0", "center"),
    7: ("#,##0.0", "center"), 8: ("#,##0.0", "center"), 9: (None, "center"),
    10: (None, "center"), 11: (None, "center"),
}

for i, c in enumerate(companies):
    row = 4 + i
    vals = [
        i + 1, c["name"], c["mtd"], c["active_days"], TOTAL_DAYS,
        c["active_pct"], c["avg_active"], c["avg_calendar"],
        c["max_streak"], str(c["first_call"]) if c["first_call"] else "",
        c["status"],
    ]
    for col, v in enumerate(vals, 1):
        cell = ws2.cell(row=row, column=col, value=v)
        fmt, align = fmts.get(col, (None, "center"))
        sc(cell, align=align, fmt=fmt)

    # Status color
    ws2.cell(row=row, column=11).fill = status_fills.get(c["status"], alt_fill)

    if i % 2 == 1:
        for col in range(1, len(hdrs) + 1):
            if col != 11:
                ws2.cell(row=row, column=col).fill = alt_fill

# Status summary block
sr = 4 + len(companies) + 2
ws2.cell(row=sr, column=1, value="Status Summary").font = Font(name="Calibri", bold=True, size=12, color=NAVY)
from collections import Counter
status_counts = Counter(c["status"] for c in companies)
for i, s in enumerate(ALL_STATUSES):
    r = sr + 1 + i
    ws2.cell(row=r, column=1, value=s).fill = status_fills[s]
    ws2.cell(row=r, column=1).font = Font(name="Calibri", bold=True, size=10)
    ws2.cell(row=r, column=2, value=f"{status_counts.get(s, 0)} companies")
    calls_in_status = sum(c["mtd"] for c in companies if c["status"] == s)
    ws2.cell(row=r, column=3, value=f"{calls_in_status:,} calls")
    pct = round(calls_in_status / total_calls * 100, 1) if total_calls else 0
    ws2.cell(row=r, column=4, value=f"{pct}% of volume")

ws2.column_dimensions["A"].width = 6
ws2.column_dimensions["B"].width = 42
for col, w in zip(range(3, 12), [12, 12, 12, 10, 14, 15, 14, 12, 16]):
    ws2.column_dimensions[get_column_letter(col)].width = w

ws2.freeze_panes = "B4"
ws2.auto_filter.ref = f"A3:{get_column_letter(len(hdrs))}{3 + len(companies)}"

# ═══════════════════════════════════════════════
# TAB 3: Top 20 Chart
# ═══════════════════════════════════════════════
ws3 = wb.create_sheet("Top 20 Chart")
ws3.sheet_properties.tabColor = "059669"

ws3.merge_cells("A1:D1")
ws3["A1"].value = "Top 20 by Avg Calls per Active Day \u2014 April 2026"
ws3["A1"].font = Font(name="Calibri", bold=True, size=13, color=NAVY)

ch = ["Company", "Avg/Active Day", "Active Days", "MTD Calls"]
for col, h in enumerate(ch, 1):
    ws3.cell(row=3, column=col, value=h)
header_row(ws3, 3, len(ch))

top20 = sorted(companies, key=lambda c: -c["avg_active"])[:20]
for i, c in enumerate(top20):
    row = 4 + i
    sc(ws3.cell(row=row, column=1, value=c["name"]), align="left")
    sc(ws3.cell(row=row, column=2, value=c["avg_active"]), fmt="#,##0.0")
    sc(ws3.cell(row=row, column=3, value=c["active_days"]))
    sc(ws3.cell(row=row, column=4, value=c["mtd"]), fmt="#,##0")

chart = BarChart()
chart.type = "bar"
chart.style = 10
chart.title = "Top 20 \u2014 Avg Calls per Active Day"
chart.y_axis.title = "Avg Calls/Active Day"
chart.width = 28
chart.height = 16
data_ref = Reference(ws3, min_col=2, min_row=3, max_row=23)
cats_ref = Reference(ws3, min_col=1, min_row=4, max_row=23)
chart.add_data(data_ref, titles_from_data=True)
chart.set_categories(cats_ref)
chart.series[0].graphicalProperties.solidFill = TEAL
ws3.add_chart(chart, "F3")

ws3.column_dimensions["A"].width = 42
ws3.column_dimensions["B"].width = 16
ws3.column_dimensions["C"].width = 12
ws3.column_dimensions["D"].width = 12

# ── Save ──
os.makedirs(OUT_DIR, exist_ok=True)
wb.save(OUT_FILE)
print(f"\nSaved: {OUT_FILE}")
print(f"Tabs:  Daily Call Matrix | Activation Status | Top 20 Chart")
print(f"Companies: {len(companies)}, Calls: {total_calls:,}, Days: {TOTAL_DAYS}")

# Status breakdown
for s in ALL_STATUSES:
    n = status_counts.get(s, 0)
    print(f"  {s}: {n} companies")
