"""HOAi Daily Report — Report Generator.

Reads a daily data JSON payload and generates:
  1. Excel workbook (1 tab: Voice_Usage Trends)
     — Voice revenue management workbook with contract data, monthly history,
       utilization status, adoption signals, and quality metrics per company.
  2. PDF report (4 pages, branded, landscape)

Voice data is MTD-scoped (1st → report date).  SMS/Webchat remain single-day.

Usage:
    python daily-reports/generate-daily-report.py                          # Latest JSON in data/
    python daily-reports/generate-daily-report.py --date 2026-04-04       # Specific date
    python daily-reports/generate-daily-report.py --json path/to/file.json
"""

import argparse
import glob
import json
import os
import sys
from datetime import datetime

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(SCRIPT_DIR, 'data')
OUTPUT_DIR = os.path.join(SCRIPT_DIR, 'output')
TEMPLATE_PATH = os.path.join(SCRIPT_DIR, 'templates', 'daily-report-template.html')
CONFIG_PATH = os.path.join(SCRIPT_DIR, 'daily-report-config.json')

os.makedirs(OUTPUT_DIR, exist_ok=True)

with open(CONFIG_PATH) as f:
    CONFIG = json.load(f)

# ── Dependency checks ──────────────────────────────────────────────────────

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)

try:
    from jinja2 import Environment, FileSystemLoader
except ImportError:
    print("ERROR: jinja2 not installed. Run: pip install jinja2")
    sys.exit(1)

# ── Style constants (matching financial-dashboard) ─────────────────────────

NAVY = '1B2A4A'
BLUE = '2563EB'
TEAL = '0D9488'
GREEN = '059669'
AMBER = 'D97706'
RED = 'DC2626'
GRAY = '6B7280'
DARK = '111827'
WHITE = 'FFFFFF'
SURFACE = 'F3F4F6'

HEADER_FILL = PatternFill('solid', fgColor=NAVY)
HEADER_FONT = Font(name='Calibri', bold=True, color=WHITE, size=10)
BLUE_HEADER_FILL = PatternFill('solid', fgColor=BLUE)
TEAL_HEADER_FILL = PatternFill('solid', fgColor=TEAL)
GREEN_HEADER_FILL = PatternFill('solid', fgColor=GREEN)
BODY_FONT = Font(name='Calibri', color=DARK, size=10)
GOOD_FONT = Font(name='Calibri', color=GREEN, size=10, bold=True)
WARN_FONT = Font(name='Calibri', color=AMBER, size=10, bold=True)
BAD_FONT = Font(name='Calibri', color=RED, size=10, bold=True)
GOOD_FILL = PatternFill('solid', fgColor='ECFDF5')
BAD_FILL = PatternFill('solid', fgColor='FEF2F2')
TOTAL_FILL = PatternFill('solid', fgColor=SURFACE)
TOTAL_FONT = Font(name='Calibri', bold=True, color=DARK, size=10)
MUTED_FONT = Font(name='Calibri', color=GRAY, size=9)
THIN_BORDER = Border(bottom=Side(style='thin', color='E5E7EB'))
THICK_BORDER = Border(
    top=Side(style='medium', color=NAVY),
    bottom=Side(style='medium', color=NAVY),
)

FMT_PCT = '0.0%'
FMT_USD = '#,##0.00'
FMT_NUM = '#,##0'

# ── Revenue workbook style constants ─────────────────────────────────────

TITLE_FONT = Font(name='Calibri', bold=True, color=DARK, size=14)
SECTION_FONT = Font(name='Calibri', bold=True, color='4338CA', size=12)
SUBHEADER_FILL = PatternFill('solid', fgColor='EEF2FF')
RISK_FILL = PatternFill('solid', fgColor='FEF2F2')
WATCH_FILL = PatternFill('solid', fgColor='FFF7ED')
HEALTHY_FILL = PatternFill('solid', fgColor='ECFDF5')
STATUS_FILLS = {'At Risk': RISK_FILL, 'Watch': WATCH_FILL, 'Healthy': HEALTHY_FILL}
STATUS_FONTS = {
    'At Risk': Font(name='Calibri', color=RED, size=10, bold=True),
    'Watch': Font(name='Calibri', color=AMBER, size=10, bold=True),
    'Healthy': Font(name='Calibri', color=GREEN, size=10, bold=True),
}
KPI_LABEL_FONT = Font(name='Calibri', color=GRAY, size=9)
KPI_VALUE_FONT = Font(name='Calibri', bold=True, color=DARK, size=14)
LIFECYCLE_FONTS = {
    'Onboard': Font(name='Calibri', color=BLUE, size=10),
    'Activate': Font(name='Calibri', color=TEAL, size=10),
    'Grow': Font(name='Calibri', color=GREEN, size=10),
    'Renew': Font(name='Calibri', color=AMBER, size=10, bold=True),
}
ADOPTION_FONTS = {
    'Ready to Expand': Font(name='Calibri', color=GREEN, size=10, bold=True),
    'Growing': Font(name='Calibri', color=GREEN, size=10),
    'Stable': Font(name='Calibri', color=GRAY, size=10),
    'Declining': Font(name='Calibri', color=RED, size=10),
}
ADOPTION_FILLS = {
    'Ready to Expand': HEALTHY_FILL,
    'Declining': RISK_FILL,
}


def _safe(val, default=0):
    return val if val is not None else default


def _apply_header(ws, row, col_count, fill=HEADER_FILL):
    for c in range(1, col_count + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)


def _apply_total_row(ws, row, col_count):
    for c in range(1, col_count + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = TOTAL_FILL
        cell.font = TOTAL_FONT
        cell.border = THICK_BORDER


def _auto_width(ws, min_width=10, max_width=28):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(min(max_len + 2, max_width), min_width)


def _rate_font(value, metric, channel='voice'):
    """Return appropriate font based on benchmark thresholds."""
    benchmarks = CONFIG.get(f'{channel}_benchmarks', {})
    bench = benchmarks.get(metric, {})
    target = bench.get('target')
    warn = bench.get('warn')
    if target is None:
        return BODY_FONT
    if target > warn:
        if value >= target:
            return GOOD_FONT
        elif value >= warn:
            return WARN_FONT
        else:
            return BAD_FONT
    else:
        if value <= target:
            return GOOD_FONT
        elif value <= warn:
            return WARN_FONT
        else:
            return BAD_FONT


def _margin_font(value):
    if value >= 0.55:
        return GOOD_FONT
    elif value >= 0.40:
        return WARN_FONT
    return BAD_FONT


def _write_kpi(ws, row, col, label, value, fmt=None):
    """Write a 2-row vertical KPI card (label on top, value below)."""
    label_cell = ws.cell(row=row, column=col, value=label)
    label_cell.font = KPI_LABEL_FONT
    label_cell.alignment = Alignment(horizontal='center')
    value_cell = ws.cell(row=row + 1, column=col, value=value)
    value_cell.font = KPI_VALUE_FONT
    value_cell.alignment = Alignment(horizontal='center')
    if fmt:
        value_cell.number_format = fmt


def _write_section_header(ws, row, start_col, end_col, title):
    """Write a merged section header with navy fill."""
    ws.merge_cells(start_row=row, start_column=start_col,
                   end_row=row, end_column=end_col)
    cell = ws.cell(row=row, column=start_col, value=title)
    cell.fill = HEADER_FILL
    cell.font = Font(name='Calibri', bold=True, color=WHITE, size=11)
    cell.alignment = Alignment(horizontal='left', vertical='center')


# ═══════════════════════════════════════════════════════════════════════════
# EXCEL GENERATION — Channel-Based Revenue Management Workbook
# ═══════════════════════════════════════════════════════════════════════════

def generate_excel(data, output_path):
    """Generate Voice-only Excel workbook from daily report JSON.

    Single tab: Voice_Usage Trends — contract data, monthly history,
    utilization status, adoption signals, and quality metrics per company.
    """
    wb = Workbook()

    # ── Shared data extraction ───────────────────────────────────────
    voice = data.get('voice', {})
    trends = data.get('historical_trends', {})
    per_company_trends = trends.get('per_company', {})
    platform = data.get('platform', {})
    report_date = data.get('report_date', '')

    monthly_labels = trends.get('monthly_labels', ['M-3', 'M-2', 'M-1', 'Proj'])
    weekly_labels = trends.get('weekly_labels', ['W-3', 'W-2', 'W-1', 'W0'])

    # Use first 3 monthly labels for history columns (exclude current month projection)
    ml = monthly_labels[:3] if len(monthly_labels) >= 3 else monthly_labels

    plat_trends = trends.get('platform', {})
    contracts = CONFIG.get('per_customer_contracts', {})
    util_thresh = CONFIG.get('utilization_thresholds', {})

    # ── Helper closures ──────────────────────────────────────────────

    def _contract(company):
        """Return contract dict or empty dict for a company."""
        return contracts.get(company, {})

    def _included(company):
        """Return included calls from contract, or None."""
        c = _contract(company)
        val = c.get('included_calls')
        if val is not None and val != 'n/a':
            return val
        return None

    def _rate_per_call(company):
        c = _contract(company)
        val = c.get('rate_per_call')
        if c.get('pricing_model') == 'per_door':
            return 'Per Door Rate'
        return val

    def _overage_per_call(company):
        c = _contract(company)
        val = c.get('overage_per_call')
        if c.get('pricing_model') == 'per_door':
            return 'Per Door Rate'
        return val

    def _utilization_status(proj_eom, included):
        """Compute utilization status: Under / Watch / On Track / Over / —"""
        if included is None or included <= 0:
            return '\u2014'
        ratio = proj_eom / included if included > 0 else 0
        under = util_thresh.get('under_pct', 0.75)
        watch_lo = util_thresh.get('watch_low_pct', 0.85)
        watch_hi = util_thresh.get('watch_high_pct', 1.15)
        over = util_thresh.get('over_pct', 1.30)
        if ratio < under:
            return 'Under'
        if ratio < watch_lo:
            return 'Watch'
        if ratio <= watch_hi:
            return 'On Track'
        if ratio <= over:
            return 'Watch'
        return 'Over'

    def _adoption_signal(t_data, forecast=None):
        """Compute adoption signal from trend data."""
        if not t_data:
            return '\u2014'
        mv = [v or 0 for v in t_data.get('monthly_values', [0, 0, 0])[:3]]
        if forecast is None:
            forecast = t_data.get('forecast_vs_avg', 'New')

        # New: first 2 months have no data
        if mv[0] == 0 and mv[1] == 0:
            return 'New'
        if forecast == 'New':
            return 'New'

        # 3-month trajectory
        m1, m2, m3 = mv[0], mv[1], mv[2]
        if m1 > 0:
            growth_3m = (m3 - m1) / m1
        elif m3 > 0:
            growth_3m = 1.0
        else:
            growth_3m = 0

        mom = t_data.get('mom_pct_change', 0)

        # Growing: 3-month trend up and not collapsing this month
        if growth_3m > 0.15 and mom > -60:
            return 'Growing'
        # Declining: negative trajectory or very negative current month
        if growth_3m < -0.20 or mom < -60:
            return 'Declining'
        # Edge case: flat 3M but current month strongly negative
        if mom < -40:
            return 'Declining'
        return 'Stable'

    def _status_font(status):
        if status == 'Under':
            return BAD_FONT
        if status == 'Over':
            return Font(name='Calibri', color=BLUE, size=10, bold=True)
        if status == 'Watch':
            return WARN_FONT
        if status == 'On Track':
            return GOOD_FONT
        return MUTED_FONT

    def _write_meta_row(ws, report_date_str):
        """Write metadata row 1 with report date."""
        ws.cell(row=1, column=1, value='Report Date:').font = BODY_FONT
        try:
            from datetime import date as _date
            rd = datetime.strptime(report_date_str, '%Y-%m-%d')
            # Excel date serial
            ws.cell(row=1, column=2, value=rd).number_format = 'YYYY-MM-DD'
        except Exception:
            ws.cell(row=1, column=2, value=report_date_str)

    # ══════════════════════════════════════════════════════════════════
    # TAB 1: VOICE_USAGE TRENDS
    # ══════════════════════════════════════════════════════════════════
    ws = wb.active
    ws.title = 'Voice_Usage Trends'

    _write_meta_row(ws, report_date)

    # Headers (row 2)
    voice_headers = [
        'Company', 'Channel', 'Go-Live Date',
        'Included Calls', '$ / Call', '$ / Overage Call',
        ml[0] if len(ml) > 0 else 'M-3',
        ml[1] if len(ml) > 1 else 'M-2',
        ml[2] if len(ml) > 2 else 'M-1',
        '3M Avg', 'MTD', 'Proj EOM', 'MoM %',
        'Status', 'Proj # of Overage Calls',
        'Base Bill $', 'Overage Bill $', 'Total Proj Bill $',
        'Adoption Signal',
        'Deflection Rate', 'Resolution %', 'CSAT', 'Error %', 'ID Rate %',
    ]
    for c, h in enumerate(voice_headers, 1):
        ws.cell(row=2, column=c, value=h)
    _apply_header(ws, 2, len(voice_headers))
    ws.freeze_panes = 'A3'

    # Collect all companies with voice trend data (historical + today)
    voice_companies = set()
    for co, ct in per_company_trends.items():
        if ct.get('voice'):
            voice_companies.add(co)
    for co in voice.keys():
        voice_companies.add(co)
    voice_companies = sorted(voice_companies)

    row = 3
    for co in voice_companies:
        t = per_company_trends.get(co, {}).get('voice') or {}
        v = voice.get(co, {})
        mv = t.get('monthly_values', [0, 0, 0, 0]) if t else [0, 0, 0, 0]

        included = _included(co)
        rate = _rate_per_call(co)
        overage = _overage_per_call(co)

        m1 = mv[0] if len(mv) > 0 else 0
        m2 = mv[1] if len(mv) > 1 else 0
        m3 = mv[2] if len(mv) > 2 else 0
        avg_3m = (m1 + m2 + m3) / 3 if any([m1, m2, m3]) else 0

        mtd = t.get('mtd_actual', 0) if t else 0
        proj_eom = t.get('projected_eom', 0) if t else 0
        mom_pct = t.get('mom_pct_change', 0) / 100 if t else 0

        status = _utilization_status(proj_eom, included)
        overage_calls = (proj_eom - included) if included is not None and included > 0 else 0

        # Billing projections
        is_per_door = _contract(co).get('pricing_model') == 'per_door'
        if is_per_door:
            base_bill = 'Per Door'
            overage_bill = 'Per Door'
            total_proj_bill = 'Per Door'
        elif included is not None and isinstance(rate, (int, float)):
            base_bill = included * rate
            ov_rate = overage if isinstance(overage, (int, float)) else 0
            overage_bill = max(0, overage_calls) * ov_rate if overage_calls > 0 else 0
            total_proj_bill = base_bill + overage_bill
        else:
            base_bill = None
            overage_bill = None
            total_proj_bill = None

        signal = _adoption_signal(t)

        # Quality metrics (from today's data, or '—' if no calls today)
        has_daily = co in voice and v.get('total_calls', 0) > 0
        defl = v.get('deflection_rate', 0) if has_daily else '\u2014'
        resol = '\u2014'
        if has_daily:
            csat_cov = v.get('csat_coverage', 0) or 0
            resolved = v.get('resolved', 0) or 0
            if csat_cov > 0:
                resol = resolved / csat_cov
            elif v.get('total_calls', 0) > 0 and resolved > 0:
                resol = resolved / v['total_calls']
        csat = v.get('avg_csat') if has_daily and v.get('avg_csat') else '\u2014'
        err_rate = v.get('error_rate_actionable', 0) if has_daily else '\u2014'
        id_rate = v.get('identified_rate', 0) if has_daily else '\u2014'

        vals = [
            co, 'Voice', None,  # Go-Live Date placeholder
            included if included is not None else ('n/a' if is_per_door else None),
            rate, overage,
            m1, m2, m3,
            round(avg_3m, 1) if avg_3m else 0,
            mtd, proj_eom,
            mom_pct,
            status,
            overage_calls,
            base_bill, overage_bill, total_proj_bill,
            signal,
            defl, resol, csat, err_rate, id_rate,
        ]

        for c, val in enumerate(vals, 1):
            cell = ws.cell(row=row, column=c, value=val)
            cell.border = THIN_BORDER

        # Format: volume columns
        for col_idx in [7, 8, 9, 11, 12]:
            ws.cell(row=row, column=col_idx).number_format = FMT_NUM
        ws.cell(row=row, column=10).number_format = '0.0'  # 3M Avg
        ws.cell(row=row, column=13).number_format = '+0.0%;-0.0%'  # MoM%
        ws.cell(row=row, column=15).number_format = '#,##0'  # Overage calls

        # Billing columns (16=Base Bill, 17=Overage Bill, 18=Total Proj Bill)
        for col_idx in [16, 17, 18]:
            cell = ws.cell(row=row, column=col_idx)
            if isinstance(cell.value, (int, float)):
                cell.number_format = FMT_USD
        # Color Total Proj Bill: amber if overage, green if on track
        tpb_cell = ws.cell(row=row, column=18)
        if isinstance(total_proj_bill, (int, float)):
            if isinstance(overage_bill, (int, float)) and overage_bill > 0:
                tpb_cell.font = WARN_FONT
            else:
                tpb_cell.font = GOOD_FONT
        elif total_proj_bill is None:
            tpb_cell.font = MUTED_FONT

        # Status color
        ws.cell(row=row, column=14).font = _status_font(status)

        # MoM color
        ws.cell(row=row, column=13).font = (
            GOOD_FONT if mom_pct > 0 else BAD_FONT if mom_pct < -0.001 else MUTED_FONT)

        # Adoption signal color (col 19)
        sig_cell = ws.cell(row=row, column=19)
        sig_cell.font = ADOPTION_FONTS.get(signal, MUTED_FONT)

        # Quality metrics formatting (cols 20-24)
        if has_daily and defl != '\u2014':
            ws.cell(row=row, column=20).number_format = FMT_PCT
            ws.cell(row=row, column=20).font = _rate_font(defl, 'deflection_rate')
        else:
            ws.cell(row=row, column=20).font = MUTED_FONT

        if resol != '\u2014':
            ws.cell(row=row, column=21).number_format = FMT_PCT
        else:
            ws.cell(row=row, column=21).font = MUTED_FONT

        if csat != '\u2014':
            ws.cell(row=row, column=22).number_format = '0.00'
        else:
            ws.cell(row=row, column=22).font = MUTED_FONT

        if err_rate != '\u2014':
            ws.cell(row=row, column=23).number_format = FMT_PCT
        else:
            ws.cell(row=row, column=23).font = MUTED_FONT

        if id_rate != '\u2014':
            ws.cell(row=row, column=24).number_format = FMT_PCT
        else:
            ws.cell(row=row, column=24).font = MUTED_FONT

        row += 1

    # ── Platform total rows ──────────────────────────────────────────
    pt_voice = plat_trends.get('voice', {})
    if pt_voice:
        pmv = pt_voice.get('monthly_values', [0, 0, 0, 0])

        # Compute platform billing totals by summing per-company bills
        plat_base_bill = 0
        plat_overage_bill = 0
        for co_name in voice_companies:
            c_data = contracts.get(co_name, {})
            if not isinstance(c_data, dict):
                continue
            if c_data.get('pricing_model') == 'per_door':
                continue
            inc = c_data.get('included_calls')
            rpc = c_data.get('rate_per_call')
            opc = c_data.get('overage_per_call')
            if inc is not None and isinstance(rpc, (int, float)):
                plat_base_bill += inc * rpc
                co_t = per_company_trends.get(co_name, {}).get('voice') or {}
                co_proj = co_t.get('projected_eom', 0) if co_t else 0
                co_ov = max(0, co_proj - inc) if inc > 0 else 0
                if co_ov > 0 and isinstance(opc, (int, float)):
                    plat_overage_bill += co_ov * opc
        plat_total_bill = plat_base_bill + plat_overage_bill

        vals_plat = [
            'PLATFORM TOTAL', 'Voice', None, None, None, None,
            pmv[0] if len(pmv) > 0 else 0,
            pmv[1] if len(pmv) > 1 else 0,
            pmv[2] if len(pmv) > 2 else 0,
            round(sum(pmv[:3]) / 3, 1) if any(pmv[:3]) else 0,
            pt_voice.get('mtd_actual', 0),
            pt_voice.get('projected_eom', 0),
            pt_voice.get('mom_pct_change', 0) / 100,
            None,
            -(sum(c_data.get('included_calls', 0) or 0
                  for c_data in contracts.values()
                  if isinstance(c_data, dict) and isinstance(c_data.get('included_calls'), (int, float)))
              - pt_voice.get('projected_eom', 0)),
            plat_base_bill, plat_overage_bill, plat_total_bill,
        ]
        for c, val in enumerate(vals_plat, 1):
            ws.cell(row=row, column=c, value=val)
        _apply_total_row(ws, row, len(voice_headers))
        for col_idx in [7, 8, 9, 11, 12]:
            ws.cell(row=row, column=col_idx).number_format = FMT_NUM
        ws.cell(row=row, column=13).number_format = '+0.0%;-0.0%'
        ws.cell(row=row, column=15).number_format = '#,##0'
        for col_idx in [16, 17, 18]:
            ws.cell(row=row, column=col_idx).number_format = FMT_USD

    _auto_width(ws, min_width=10, max_width=30)
    ws.column_dimensions['A'].width = 35  # Company names can be long

    wb.save(output_path)
    print(f"  Excel: {output_path}")


# ═══════════════════════════════════════════════════════════════════════════
# PDF GENERATION
# ═══════════════════════════════════════════════════════════════════════════

def _commafy(val):
    """Jinja2 filter: format number with commas."""
    if val is None:
        return '0'
    try:
        if isinstance(val, float):
            return f"{val:,.2f}"
        return f"{int(val):,}"
    except (ValueError, TypeError):
        return str(val)


def _pctfmt(val):
    """Jinja2 filter: format rate (0-1) as percentage string."""
    if val is None:
        return '\u2014'
    try:
        return f"{float(val) * 100:.1f}%"
    except (ValueError, TypeError):
        return str(val)


def _rate_class(val, metric, channel='voice'):
    """Jinja2 filter: return CSS class based on benchmark."""
    benchmarks = CONFIG.get(f'{channel}_benchmarks', {})
    bench = benchmarks.get(metric, {})
    target = bench.get('target')
    warn = bench.get('warn')
    if target is None or val is None:
        return ''
    try:
        val = float(val)
    except (ValueError, TypeError):
        return ''
    if target > warn:
        if val >= target:
            return 'good'
        elif val >= warn:
            return 'warn'
        return 'bad'
    else:
        if val <= target:
            return 'good'
        elif val <= warn:
            return 'warn'
        return 'bad'


def generate_pdf(data, output_path):
    """Generate 4-page landscape PDF from Jinja2 template (Voice-only).

    Falls back to saving standalone HTML if weasyprint is unavailable.
    """
    try:
        from weasyprint import HTML
        has_weasyprint = True
    except (ImportError, OSError):
        has_weasyprint = False
        print("  WARN: weasyprint unavailable (needs GTK/Pango). Generating HTML instead.")
        print("        Install GTK: https://doc.courtbouillon.org/weasyprint/stable/first_steps.html")

    env = Environment(
        loader=FileSystemLoader(os.path.join(SCRIPT_DIR, 'templates')),
        autoescape=False,
    )
    env.filters['commafy'] = _commafy
    env.filters['pctfmt'] = _pctfmt
    env.filters['rate_class'] = _rate_class

    template = env.get_template('daily-report-template.html')

    # Prepare template context with safe defaults for missing summaries
    platform = data.get('platform', {})
    _voice_defaults = {
        'total_calls': 0, 'deflection_rate': 0, 'transfer_rate': 0,
        'error_rate_actionable': 0, 'active_companies': 0,
        'hours_saved': 0, 'dollar_saved': 0,
    }
    voice_plat = {**_voice_defaults, **platform.get('voice_summary', {})}

    generated_at = data.get('generated_at', '')
    try:
        generated_at_short = datetime.fromisoformat(generated_at).strftime('%b %d, %Y %I:%M %p')
    except Exception:
        generated_at_short = generated_at

    # MTD context for banner subtitle
    report_date_str = data.get('report_date', '')
    try:
        rd = datetime.strptime(report_date_str, '%Y-%m-%d')
        mtd_start_display = rd.replace(day=1).strftime('%b %d')
        days_elapsed = rd.day
    except Exception:
        mtd_start_display = ''
        days_elapsed = 0

    html_content = template.render(
        report_date=report_date_str,
        generated_at_short=generated_at_short,
        mtd_start_display=mtd_start_display,
        days_elapsed=days_elapsed,
        platform=platform,
        voice_plat=voice_plat,
        voice=data.get('voice', {}),
        revenue_intelligence=data.get('revenue_intelligence', []),
        historical_trends=data.get('historical_trends', {}),
        contracts=CONFIG.get('per_customer_contracts', {}),
    )

    if has_weasyprint:
        HTML(string=html_content).write_pdf(output_path)
        print(f"  PDF:   {output_path}")
    else:
        # Save as standalone HTML (can be opened in browser, printed to PDF)
        html_path = output_path.replace('.pdf', '.html')
        with open(html_path, 'w', encoding='utf-8') as f:
            f.write(html_content)
        print(f"  HTML:  {html_path} (open in browser -> Print -> Save as PDF)")


# ═══════════════════════════════════════════════════════════════════════════
# PER-COMPANY OUTREACH PAGES
# ═══════════════════════════════════════════════════════════════════════════

OUTREACH_TEMPLATE_PATH = os.path.join(SCRIPT_DIR, 'templates', 'outreach-template.html')


def _adoption_signal(trend_data):
    """Classify volume trend as Accelerating / Stable / Declining."""
    mv = (trend_data.get('monthly_values') or [0, 0, 0])[:3]
    if len(mv) < 2:
        return 'Stable'
    mom = trend_data.get('mom_pct_change')
    if mom is not None:
        if mom > 10:
            return 'Accelerating'
        elif mom < -10:
            return 'Declining'
    return 'Stable'


def _extract_top_transfer_topics(topics, limit=3):
    """Return top transfer topics where transfer_rate > 40%, sorted by transferred desc."""
    candidates = [t for t in topics
                  if t.get('transfer_rate_pct', 0) > 40
                  and t.get('topic') != 'Incomplete & Non-Engagement'
                  and t.get('transferred', 0) > 2]
    candidates.sort(key=lambda t: t.get('transferred', 0), reverse=True)
    result = []
    for t in candidates[:limit]:
        est_deflectable = int(t['transferred'] * 0.15)  # 15pp improvement estimate
        result.append({
            'topic': t['topic'],
            'count': t.get('count', 0),
            'transferred': t.get('transferred', 0),
            'transfer_rate': t.get('transfer_rate_pct', 0),
            'deflection_rate': t.get('deflection_rate_pct', 0),
            'est_deflectable': est_deflectable,
            'recommendation': _topic_recommendation(t['topic']),
        })
    return result


def _topic_recommendation(topic_name):
    """Map topic to specific AOP improvement recommendation."""
    recs = {
        'Account & Identity': 'Add identity verification action items to the AOP — enable lookup by address, name, or account number to resolve without transfer.',
        'Billing & Payments': 'Ensure the agent has access to payment status, balance lookups, and payment portal links. Add a "make a payment" action item.',
        'Transfer Requests': 'Review transfer routing rules — some "transfer requests" may be resolvable if the agent has the right tools or extensions.',
        'Maintenance & Violations': 'Add work order submission and violation status lookup as action items. These are high-deflection opportunities.',
        'Community & Amenity': 'Add community info, amenity hours, and reservation links to the agent knowledge base.',
        'Governance & Meetings': 'Add board meeting schedules, document links, and ACC submission forms to the AOP.',
        'Emergency & Urgent': 'Verify after-hours emergency routing is configured correctly. These calls should transfer to on-call staff.',
        'General Inquiry': 'Broaden the agent knowledge base — general inquiries often indicate missing FAQ content.',
        'Complaints & Escalation': 'Review complaint handling flow. Consider adding an empathy response + ticket creation action item.',
        'Insurance & Legal': 'Add insurance certificate request and legal contact information to the agent.',
    }
    return recs.get(topic_name, f'Review AOP for {topic_name} — adding relevant action items could reduce transfers.')


def _get_unidentified_pct(caller_types):
    """Return percentage of unidentified callers."""
    for ct in caller_types:
        if 'unidentified' in ct.get('caller_type', '').lower():
            return ct.get('pct', 0)
    return 0


def _get_lowest_csat_dim(dims):
    """Find the CSAT dimension with the lowest score. Returns (name, score) or None."""
    if not dims:
        return None
    label_map = {
        'task_completion': 'Task Completion',
        'caller_identification': 'Caller Identification',
        'call_forwarding': 'Call Forwarding',
        'distortion': 'Audio Quality',
        'repetition': 'Repetition',
        'interruptions': 'Interruptions',
        'technical_failure': 'Technical Reliability',
    }
    lowest_key, lowest_val = None, 999
    for k, v in dims.items():
        if v is not None and v < lowest_val:
            lowest_key, lowest_val = k, v
    if lowest_key is None:
        return None
    return {'key': lowest_key, 'label': label_map.get(lowest_key, lowest_key), 'score': round(lowest_val, 1)}


def _csat_dim_recommendation(dim_key):
    """Return specific recommendation for a low CSAT dimension."""
    recs = {
        'task_completion': 'Task completion is low — the agent may lack sufficient knowledge or tools to resolve requests. Review the AOP for missing action items.',
        'caller_identification': 'Caller identification is low — homeowners may feel unrecognized. Ensure the agent asks for identifying info early in the call.',
        'call_forwarding': 'Call forwarding experience is low — transfers may feel abrupt. Add warm transfer context or explain why the transfer is needed.',
        'distortion': 'Audio quality scores are low — check phone line quality and Vapi voice settings.',
        'repetition': 'Repetition score is low — the agent may be asking the same questions multiple times. Review conversation flow for loops.',
        'interruptions': 'Interruption score is low — the agent may be cutting off callers. Check voice activity detection settings.',
        'technical_failure': 'Technical reliability is low — check for dropped calls or system errors in the error breakdown.',
    }
    return recs.get(dim_key, 'Review this CSAT dimension for improvement opportunities.')


def _calc_time_saved(duration_by_outcome):
    """Compute time saved per deflection (transferred avg - deflected avg seconds)."""
    deflected_avg = None
    transferred_avg = None
    deflected_count = 0
    for d in duration_by_outcome:
        if d.get('outcome') == 'deflected':
            deflected_avg = d.get('avg_seconds')
            deflected_count = d.get('count', 0)
        elif d.get('outcome') == 'transferred':
            transferred_avg = d.get('avg_seconds')
    if deflected_avg is not None and transferred_avg is not None:
        return {
            'deflected_avg': round(deflected_avg, 1),
            'transferred_avg': round(transferred_avg, 1),
            'saved_per_defl': round(transferred_avg - deflected_avg, 1),
            'deflected_count': deflected_count,
            'total_hours_saved': round(deflected_count * deflected_avg / 3600, 1),
        }
    return None


def _build_billing_analysis(contract, trend_data, voice_packages):
    """Build detailed billing/usage analysis with upgrade economics."""
    included = contract.get('included_calls')
    overage_rate = contract.get('overage_per_call') or 0
    rate_per_call = contract.get('rate_per_call') or 0

    if not included or included <= 0:
        return None

    mtd = trend_data.get('mtd_actual', 0)
    proj = trend_data.get('projected_eom', 0)
    mv = (trend_data.get('monthly_values') or [])[:4]
    trailing_3m = trend_data.get('trailing_3m_avg', 0)

    # Monthly overage history
    monthly_history = []
    month_labels = ['This Month (Proj.)', 'Last Month', '2 Months Ago', '3 Months Ago']
    values_to_check = [proj] + mv[1:] if mv else [proj]
    overage_months = 0
    total_overage_calls = 0
    total_overage_cost = 0.0

    for i, val in enumerate(values_to_check[:4]):
        is_over = val > included
        if is_over:
            overage_months += 1
        ovg_calls = max(0, val - included)
        ovg_cost = ovg_calls * overage_rate
        total_overage_calls += ovg_calls
        total_overage_cost += ovg_cost
        monthly_history.append({
            'label': month_labels[i] if i < len(month_labels) else f'M-{i}',
            'volume': f'{val:,}',
            'included': f'{included:,}',
            'overage_calls': ovg_calls,
            'overage_cost': f'${ovg_cost:,.2f}',
            'is_over': is_over,
            'is_projected': i == 0,
        })

    # Current effective monthly cost (base + avg overage)
    base_monthly = included * rate_per_call if rate_per_call else 0
    avg_monthly_overage = total_overage_cost / max(len(values_to_check), 1)
    current_effective = base_monthly + avg_monthly_overage

    # Find optimal tier
    tier_list = sorted(voice_packages.items(), key=lambda x: x[1]['included_calls'])
    upgrade_options = []
    for tier_name, tier in tier_list:
        if tier['included_calls'] <= included:
            continue  # Skip smaller tiers
        # Calculate cost at this tier
        tier_overage_calls = max(0, proj - tier['included_calls'])
        tier_overage_cost = tier_overage_calls * tier.get('overage_per_call', 0)
        tier_total = tier['monthly_price'] + tier_overage_cost

        # Savings vs current effective cost
        monthly_savings = current_effective - tier_total
        annual_savings = monthly_savings * 12

        upgrade_options.append({
            'tier': tier_name,
            'included': f'{tier["included_calls"]:,}',
            'monthly_price': f'${tier["monthly_price"]:,}',
            'proj_overage_at_tier': tier_overage_calls,
            'proj_overage_cost': f'${tier_overage_cost:,.2f}',
            'total_at_tier': f'${tier_total:,.2f}',
            'monthly_savings': f'${monthly_savings:,.2f}',
            'annual_savings': f'${annual_savings:,.2f}',
            'saves_money': monthly_savings > 0,
            'covers_projected': tier['included_calls'] >= proj,
        })

    # Determine recommendation strength
    is_chronic = overage_months >= 3
    is_recurring = overage_months >= 2
    best_upgrade = next((u for u in upgrade_options if u['saves_money']), None)
    covers_upgrade = next((u for u in upgrade_options if u['covers_projected']), None)

    if is_chronic:
        urgency = 'critical'
        urgency_label = 'Chronic Overage'
        urgency_detail = f'In overage {overage_months} of last {len(values_to_check)} months — this is a recurring pattern, not a one-time spike.'
    elif is_recurring:
        urgency = 'high'
        urgency_label = 'Recurring Overage'
        urgency_detail = f'In overage {overage_months} of last {len(values_to_check)} months — a pattern is forming.'
    elif proj > included:
        urgency = 'moderate'
        urgency_label = 'Projected Overage'
        urgency_detail = f'On pace to exceed included calls this month (projected: {proj:,} vs. {included:,} included).'
    else:
        return None  # No billing issue

    return {
        'urgency': urgency,
        'urgency_label': urgency_label,
        'urgency_detail': urgency_detail,
        'included_calls': f'{included:,}',
        'overage_rate': f'${overage_rate:.2f}',
        'mtd_actual': f'{mtd:,}',
        'proj_eom': f'{proj:,}',
        'overage_months': overage_months,
        'total_months': len(values_to_check),
        'total_overage_cost': f'${total_overage_cost:,.2f}',
        'avg_monthly_overage': f'${avg_monthly_overage:,.2f}',
        'current_effective': f'${current_effective:,.2f}',
        'monthly_history': monthly_history,
        'upgrade_options': upgrade_options,
        'best_upgrade': best_upgrade,
        'covers_upgrade': covers_upgrade,
        'is_chronic': is_chronic,
        'is_recurring': is_recurring,
        'trailing_3m_avg': f'{trailing_3m:,.0f}' if trailing_3m else None,
    }


def _build_recommendations(usage_status, adoption, voice_data, trend_data, alerts, cross_channel):
    """Generate outreach recommendations based on usage tag + volume trend."""
    recs = []
    mtd = trend_data.get('mtd_actual', 0)
    proj = trend_data.get('projected_eom', 0)
    mom = trend_data.get('mom_pct_change')
    defl = voice_data.get('deflection_rate')
    xfer = voice_data.get('transfer_rate')
    channels = cross_channel.get('channels', [])
    white_space = cross_channel.get('white_space', '')

    # Usage-specific recommendations
    if usage_status == 'Overage':
        recs.append({
            'title': 'Package Upgrade Discussion',
            'detail': f'Currently projected at {proj:,} calls this month, exceeding their contract. '
                      f'Schedule a call to discuss upgrading to a higher-tier package that accommodates '
                      f'their growing volume and reduces overage costs.'
        })
    elif usage_status == 'Proj. Overage':
        recs.append({
            'title': 'Proactive Overage Warning',
            'detail': f'On pace to exceed included calls by end of month (projected: {proj:,}). '
                      f'Reach out proactively to discuss usage optimization or package adjustment '
                      f'before overage charges kick in.'
        })
    elif usage_status == 'On Track':
        recs.append({
            'title': 'Value Reinforcement',
            'detail': f'Usage is healthy at {mtd:,} calls MTD. Reinforce the value being delivered — '
                      f'deflection rate of {_fmt_pct(defl)}, saving staff time on routine calls.'
        })
    elif usage_status == 'Under-Use':
        recs.append({
            'title': 'Engagement & Activation',
            'detail': f'Only {mtd:,} calls MTD — well below contract capacity. Review call routing '
                      f'configuration and ensure the HOAi number is prominently listed. '
                      f'Consider a homeowner communication campaign to drive awareness.'
        })
    else:  # No Contract
        recs.append({
            'title': 'Contract Opportunity',
            'detail': f'This company has {mtd:,} calls MTD with no formal contract. '
                      f'Use volume data to propose an appropriate package tier.'
        })

    # Adoption-specific overlay
    if adoption == 'Declining':
        recs.append({
            'title': 'Address Declining Volume',
            'detail': f'Volume is trending down (MoM: {_fmt_pct_raw(mom)}). '
                      f'Investigate if there are routing changes, staffing shifts, or dissatisfaction '
                      f'that may be causing the drop. A retention-focused check-in is recommended.'
        })
    elif adoption == 'Accelerating':
        recs.append({
            'title': 'Capitalize on Growth',
            'detail': f'Volume is accelerating (MoM: +{_fmt_pct_raw(mom)}). '
                      f'This is a great time to discuss expanded capabilities or upsell SMS/Webchat.'
        })

    # Cross-sell opportunities
    if 'S' in white_space:
        recs.append({
            'title': 'SMS Cross-Sell Opportunity',
            'detail': 'This customer has Voice but no SMS. With their call volume, '
                      'SMS broadcast and two-way messaging could deflect additional routine inquiries.'
        })
    if 'W' in white_space:
        recs.append({
            'title': 'Webchat Cross-Sell Opportunity',
            'detail': 'No Webchat channel active. Adding Webchat can capture after-hours '
                      'and younger-demographic homeowner interactions.'
        })

    # Alert-driven recommendations
    if any(a.get('metric', '').lower().startswith('transfer') for a in alerts):
        recs.append({
            'title': 'High Transfer Rate — Review Routing',
            'detail': f'Transfer rate is elevated at {_fmt_pct(xfer)}. '
                      f'Review the AOP for missing action items or knowledge gaps that could '
                      f'reduce unnecessary transfers to staff.'
        })

    return recs


def _fmt_pct(v):
    if v is None:
        return '—'
    return f'{v * 100:.1f}%'


def _fmt_pct_raw(v):
    if v is None:
        return '—'
    return f'{v:.1f}%'


def _usage_badge(status):
    mapping = {
        'Overage': 'overage',
        'Proj. Overage': 'proj-overage',
        'On Track': 'on-track',
        'Under-Use': 'under-use',
    }
    return mapping.get(status, 'no-contract')


def _compute_usage_status(mtd_actual, proj_eom, included_calls):
    if included_calls is None or included_calls <= 0:
        return None
    if mtd_actual > included_calls:
        return 'Overage'
    if proj_eom > included_calls:
        return 'Proj. Overage'
    if proj_eom >= included_calls * 0.75:
        return 'On Track'
    return 'Under-Use'


def generate_outreach_pages(data, output_base):
    """Generate per-company outreach HTML pages."""
    if not os.path.exists(OUTREACH_TEMPLATE_PATH):
        print("  WARN: outreach-template.html not found, skipping outreach pages.")
        return 0

    env = Environment(loader=FileSystemLoader(os.path.dirname(OUTREACH_TEMPLATE_PATH)))
    template = env.get_template(os.path.basename(OUTREACH_TEMPLATE_PATH))

    report_date = data.get('report_date', 'unknown')
    voice = data.get('voice', {})
    trends_raw = data.get('historical_trends', {})
    trends = trends_raw.get('per_company', {}) if isinstance(trends_raw, dict) else {}
    alerts_all = data.get('alerts', [])
    cross_channel = data.get('cross_channel', {})
    contracts = CONFIG.get('per_customer_contracts', {})
    repeat_callers = data.get('repeat_callers', [])
    platform = data.get('platform', {})
    voice_summary = platform.get('voice_summary', {})
    book_avg_defl = voice_summary.get('deflection_rate')
    days_elapsed = voice_summary.get('days_elapsed', 1)

    # Determine qualifying companies
    alert_companies = {a['company'] for a in alerts_all if 'company' in a}
    repeat_companies = {r['company_name'] for r in repeat_callers if r.get('calls_today', 0) >= 3}

    count = 0
    all_companies = set(voice.keys()) | set(trends.keys())

    for company in sorted(all_companies):
        vd = voice.get(company, {})
        vt = (trends.get(company) or {}).get('voice') or {}
        cc = cross_channel.get(company, {})
        company_alerts = [a for a in alerts_all if a.get('company') == company]
        contract = contracts.get(company, {})

        mtd_actual = vt.get('mtd_actual', vd.get('total_calls', 0))
        proj_eom = vt.get('projected_eom', 0)
        included_calls = contract.get('included_calls')
        usage_status = _compute_usage_status(mtd_actual, proj_eom, included_calls)
        adoption = _adoption_signal(vt)

        # Qualify: must have alerts, bad status, low health, repeats, or declining
        has_alerts = company in alert_companies
        has_bad_status = usage_status in ('Overage', 'Proj. Overage')
        has_repeats = company in repeat_companies
        is_declining = adoption == 'Declining' and mtd_actual > 50
        is_low_volume = usage_status == 'Under-Use'

        if not (has_alerts or has_bad_status or has_repeats or is_declining or is_low_volume):
            continue

        # Build recommendations
        recs = _build_recommendations(
            usage_status or 'No Contract', adoption, vd, vt, company_alerts, cc
        )

        # Metric formatting
        defl_rate = vd.get('deflection_rate')
        defl_color = 'var(--green)' if defl_rate and defl_rate >= 0.5 else (
            'var(--amber)' if defl_rate and defl_rate >= 0.3 else 'var(--red)'
        )
        mom_pct = vt.get('mom_pct_change')
        mom_color = 'var(--green)' if mom_pct and mom_pct > 0 else (
            'var(--red)' if mom_pct and mom_pct < 0 else 'var(--gray)'
        )

        # Usage bar
        usage_pct = min(100, (proj_eom / included_calls * 100)) if included_calls else 0
        threshold_pct = min(100, (included_calls / max(proj_eom, included_calls, 1) * 100)) if included_calls else 100
        usage_bar_color = 'red' if usage_status == 'Overage' else (
            'amber' if usage_status == 'Proj. Overage' else 'green'
        )
        overage_calls = max(0, mtd_actual - included_calls) if included_calls else 0
        overage_cost = f'${overage_calls * 0.15:.2f}' if overage_calls > 0 else ''

        # Monthly values for trend display
        monthly_values = []
        mv_raw = (vt.get('monthly_values') or [])[:3]
        month_labels = ['This Month', 'Last Month', '2 Months Ago']
        for i, val in enumerate(mv_raw):
            monthly_values.append({'label': month_labels[i] if i < len(month_labels) else f'M-{i}', 'value': f'{val:,}'})

        daily_avg = f'{mtd_actual / max(days_elapsed, 1):.0f}'

        context = {
            'company': company,
            'report_date': report_date,
            'usage_status_label': usage_status or 'No Contract',
            'usage_badge': _usage_badge(usage_status),
            'adoption_signal': adoption if adoption != 'Stable' else '',
            'mtd_actual': f'{mtd_actual:,}',
            'daily_avg': daily_avg,
            'deflection_pct': _fmt_pct(defl_rate),
            'defl_color': defl_color,
            'book_avg_defl': _fmt_pct(book_avg_defl),
            'csat': f'{vd.get("avg_csat", 0):.1f}' if vd.get('avg_csat') else '—',
            'transfer_pct': _fmt_pct(vd.get('transfer_rate')),
            'proj_eom': f'{proj_eom:,}',
            'mom_pct': _fmt_pct_raw(mom_pct) if mom_pct is not None else '—',
            'mom_color': mom_color,
            'included_calls': f'{included_calls:,}' if included_calls else None,
            'usage_pct': f'{usage_pct:.0f}',
            'threshold_pct': f'{threshold_pct:.0f}',
            'usage_bar_color': usage_bar_color,
            'overage_calls': overage_calls,
            'overage_cost': overage_cost,
            'alerts': [{'metric': a.get('metric', ''), 'value': a.get('value', ''),
                        'threshold': a.get('threshold', ''), 'severity': a.get('severity', 'medium')}
                       for a in company_alerts],
            'recommendations': recs,
            'monthly_values': monthly_values,
            # Detailed pattern data
            'top_transfer_topics': _extract_top_transfer_topics(vd.get('topics', [])),
            'hourly_pattern': vd.get('hourly_pattern', []),
            'after_hours_pct': vd.get('after_hours_pct'),
            'busiest_hour': vd.get('busiest_hour'),
            'caller_types': vd.get('caller_types', []),
            'unidentified_pct': _get_unidentified_pct(vd.get('caller_types', [])),
            'csat_dimensions': vd.get('csat_dimensions', {}),
            'lowest_csat_dim': _get_lowest_csat_dim(vd.get('csat_dimensions', {})),
            'csat_dim_rec': _csat_dim_recommendation(
                (_get_lowest_csat_dim(vd.get('csat_dimensions', {})) or {}).get('key', '')
            ),
            'duration_by_outcome': vd.get('duration_by_outcome', []),
            'time_saved': _calc_time_saved(vd.get('duration_by_outcome', [])),
            'engaged_defl': _fmt_pct(vd.get('engaged_deflection_rate')),
            'headline_defl': _fmt_pct(defl_rate),
            'defl_gap': vd.get('engaged_deflection_rate') and defl_rate and
                        abs((vd['engaged_deflection_rate'] - defl_rate) * 100) > 5,
            'revenue_intel': vd.get('revenue_intel'),
            'weekly_values': vt.get('weekly_values', []),
            # Billing & upgrade analysis
            'billing': _build_billing_analysis(contract, vt, CONFIG.get('voice_packages', {})),
        }

        html_out = template.render(**context)

        # Write to Outreach/{Company_Slug}/outreach.html
        slug = company.replace(' ', '_').replace(',', '').replace('.', '').replace("'", '')
        outreach_dir = os.path.join(output_base, 'Outreach', slug)
        os.makedirs(outreach_dir, exist_ok=True)
        outreach_path = os.path.join(outreach_dir, 'outreach.html')
        with open(outreach_path, 'w', encoding='utf-8') as f:
            f.write(html_out)
        count += 1

    return count


# ═══════════════════════════════════════════════════════════════════════════
# PER-COMPANY OVERAGE EMAILS
# ═══════════════════════════════════════════════════════════════════════════

OVERAGE_EMAIL_TEMPLATE_PATH = os.path.join(SCRIPT_DIR, 'templates', 'overage-email-template.html')


def _get_customer_tiers(company, contract, config):
    """Get available tiers for a company — custom if configured, else default ladder."""
    per_cust = config.get('per_customer_tiers', {})
    if company in per_cust:
        entry = per_cust[company]
        return entry.get('current_tier_name'), entry.get('tiers', [])

    # Fall back to default tier ladder — find current tier by included_calls match
    ladder = per_cust.get('_default_tier_ladder', [])
    included = contract.get('included_calls', 0)
    rate = contract.get('rate_per_call', 0)
    overage = contract.get('overage_per_call', 0)

    # Find closest matching tier in ladder
    current_name = 'Current'
    for t in ladder:
        if t['included_calls'] == included:
            current_name = t['name']
            break

    # Build tier list: current (from contract) + upgrades from ladder
    current_tier = {
        'name': current_name,
        'included_calls': included,
        'monthly_price': round(included * rate) if rate else 0,
        'rate_per_call': rate,
        'overage_per_call': overage,
    }
    tiers = [current_tier]
    for t in ladder:
        if t['included_calls'] > included:
            tiers.append(dict(t))
    return current_name, tiers


def _commafy(value):
    """Format an integer with commas."""
    try:
        return f'{int(value):,}'
    except (ValueError, TypeError):
        return str(value)


def _commafy_raw(value):
    """Format a number string/int with commas (no $ prefix)."""
    try:
        return f'{int(float(value)):,}'
    except (ValueError, TypeError):
        return str(value)


def generate_overage_emails(data, output_base):
    """Generate per-company overage notification emails (HTML)."""
    if not os.path.exists(OVERAGE_EMAIL_TEMPLATE_PATH):
        print("  WARN: overage-email-template.html not found, skipping overage emails.")
        return 0

    env = Environment(loader=FileSystemLoader(os.path.dirname(OVERAGE_EMAIL_TEMPLATE_PATH)))
    env.filters['commafy'] = _commafy
    env.filters['commafy_raw'] = _commafy_raw
    template = env.get_template(os.path.basename(OVERAGE_EMAIL_TEMPLATE_PATH))

    report_date = data.get('report_date', 'unknown')
    voice = data.get('voice', {})
    trends_raw = data.get('historical_trends', {})
    trends = trends_raw.get('per_company', {}) if isinstance(trends_raw, dict) else {}
    contracts = CONFIG.get('per_customer_contracts', {})
    platform = data.get('platform', {})
    voice_summary = platform.get('voice_summary', {})
    days_elapsed = voice_summary.get('days_elapsed', 1)

    # Parse report date for month labels
    try:
        rd = datetime.strptime(report_date, '%Y-%m-%d')
        month_name = rd.strftime('%B')
        today_label = rd.strftime('%B %d').replace(' 0', ' ')
        # Build month labels for historical data (current month back)
        month_labels = []
        for i in range(4):
            m = rd.month - i
            y = rd.year
            while m <= 0:
                m += 12
                y -= 1
            month_labels.append(datetime(y, m, 1).strftime('%b %Y'))
    except ValueError:
        month_name = 'This Month'
        today_label = report_date
        month_labels = ['This Month', 'Last Month', '2 Months Ago', '3 Months Ago']

    count = 0
    all_companies = set(voice.keys()) | set(trends.keys())

    for company in sorted(all_companies):
        vd = voice.get(company, {})
        vt = (trends.get(company) or {}).get('voice') or {}
        contract = contracts.get(company, {})
        if not isinstance(contract, dict):
            continue

        included = contract.get('included_calls')
        if not included or included <= 0:
            continue

        mtd_actual = vt.get('mtd_actual', vd.get('total_calls', 0))
        proj_eom = vt.get('projected_eom', 0)

        # Only generate for companies in overage or projected overage
        if proj_eom <= included and mtd_actual <= included:
            continue

        # Get tier data
        current_tier_name, tiers = _get_customer_tiers(company, contract, CONFIG)
        if not tiers:
            continue

        current_tier = tiers[0]
        upgrade_tiers_raw = tiers[1:]

        # Build monthly history (skip ramp months with < 20 calls)
        # monthly_values is chronological [oldest..newest], reverse so index 0 = current month
        mv_raw = (vt.get('monthly_values') or [])[:4]
        mv = list(reversed(mv_raw))  # Now [current, last_month, 2_months_ago, 3_months_ago]
        history_rows = []
        overage_months_count = 0
        for i, val in enumerate(mv):
            if i == 0:
                # Current month — use projected
                ovg = max(0, proj_eom - included)
                if ovg > 0:
                    overage_months_count += 1
                history_rows.append({
                    'label': f'{month_labels[0]} (proj.)',
                    'calls': proj_eom,
                    'overage': ovg,
                })
            else:
                if val < 20:
                    continue  # Skip ramp months
                ovg = max(0, val - included)
                if ovg > 0:
                    overage_months_count += 1
                history_rows.append({
                    'label': month_labels[i] if i < len(month_labels) else f'M-{i}',
                    'calls': val,
                    'overage': ovg,
                })

        # Put history in chronological order (oldest first)
        history_rows.reverse()

        # Build history narrative
        if overage_months_count >= 3:
            cum_overage = sum(max(0, v - included) * contract.get('overage_per_call', 0)
                            for v in mv[1:] if v >= 20)
            history_narrative = (
                f'This is a pattern we\'ve seen consistently since launch, with '
                f'${cum_overage:,.0f} in cumulative overage charges across the period:'
            )
        elif overage_months_count == 2:
            history_narrative = 'This is the second consecutive month your usage has exceeded your included calls:'
        elif len([r for r in history_rows if r['label'] != f'{month_labels[0]} (proj.)']) == 0:
            history_narrative = ''
            history_rows = []  # No history to show for first-month companies
        else:
            history_narrative = 'Your usage has grown steadily since going live, with this being the first month exceeding your included calls:'

        # Cost calculations for current tier
        proj_overage_calls = max(0, proj_eom - current_tier['included_calls'])
        proj_overage_cost = proj_overage_calls * current_tier.get('overage_per_call', 0)
        proj_total_current = current_tier.get('monthly_price', 0) + proj_overage_cost

        # Build comparison columns (current + upgrades)
        comparison_tiers = []
        comparison_tiers.append({
            'name': current_tier['name'],
            'is_current': True,
            'included_calls': current_tier['included_calls'],
            'rate_per_call': current_tier.get('rate_per_call', 0),
            'monthly_price': current_tier.get('monthly_price', 0),
            'overage_per_call': current_tier.get('overage_per_call', 0),
            'overage_calls': proj_overage_calls,
            'overage_cost': int(proj_overage_cost),
            'projected_total': int(proj_total_current),
            'savings': 0,
            'savings_pct': 0,
        })

        upgrade_tiers_out = []
        for ut in upgrade_tiers_raw:
            ut_overage_calls = max(0, proj_eom - ut['included_calls'])
            ut_overage_cost = ut_overage_calls * ut.get('overage_per_call', 0)
            ut_total = ut.get('monthly_price', 0) + ut_overage_cost
            savings = proj_total_current - ut_total
            savings_pct = (savings / proj_total_current * 100) if proj_total_current > 0 else 0

            tier_out = {
                'name': ut['name'],
                'is_current': False,
                'included_calls': ut['included_calls'],
                'rate_per_call': ut.get('rate_per_call', 0),
                'monthly_price': ut.get('monthly_price', 0),
                'overage_per_call': ut.get('overage_per_call', 0),
                'overage_calls': ut_overage_calls,
                'overage_cost': int(ut_overage_cost),
                'projected_total': int(ut_total),
                'savings': int(savings),
                'savings_pct': savings_pct,
            }
            comparison_tiers.append(tier_out)
            upgrade_tiers_out.append(tier_out)

        context = {
            'company': company,
            'month_name': month_name,
            'today_label': today_label,
            'report_date': report_date,
            'mtd_actual': mtd_actual,
            'projected_calls': proj_eom,
            'current_tier': current_tier,
            'projected_overage_calls': proj_overage_calls,
            'projected_overage_cost': proj_overage_cost,
            'projected_total_current': proj_total_current,
            'history_rows': history_rows,
            'history_narrative': history_narrative,
            'comparison_tiers': comparison_tiers,
            'upgrade_tiers': upgrade_tiers_out,
        }

        html_out = template.render(**context)

        slug = company.replace(' ', '_').replace(',', '').replace('.', '').replace("'", '')
        outreach_dir = os.path.join(output_base, 'Outreach', slug)
        os.makedirs(outreach_dir, exist_ok=True)
        email_path = os.path.join(outreach_dir, 'overage-email.html')
        with open(email_path, 'w', encoding='utf-8') as f:
            f.write(html_out)
        count += 1

    return count


# ═══════════════════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(description='HOAi Daily Report \u2014 Generator')
    parser.add_argument('--date', type=str, default=None,
                        help='Report date (YYYY-MM-DD). Looks for matching JSON in data/.')
    parser.add_argument('--json', type=str, default=None,
                        help='Path to specific JSON payload.')
    args = parser.parse_args()

    # Find JSON payload
    if args.json:
        json_path = args.json
    elif args.date:
        json_path = os.path.join(DATA_DIR, f'daily-report-{args.date}.json')
    else:
        # Find latest JSON in data/
        files = sorted(glob.glob(os.path.join(DATA_DIR, 'daily-report-*.json')))
        if not files:
            print("ERROR: No JSON payloads found in data/. Run fetch-daily-data.py first.")
            sys.exit(1)
        json_path = files[-1]

    if not os.path.exists(json_path):
        print(f"ERROR: JSON not found: {json_path}")
        sys.exit(1)

    print(f"HOAi Daily Report \u2014 Generating from {json_path}")
    with open(json_path) as f:
        data = json.load(f)

    report_date = data.get('report_date', 'unknown')

    # Nested folder structure: output/YYYY-MM/YYYY-MM-DD/Dashboard/
    month_dir = report_date[:7]  # "2026-04"
    dashboard_dir = os.path.join(OUTPUT_DIR, month_dir, report_date, 'Dashboard')
    os.makedirs(dashboard_dir, exist_ok=True)

    # Generate Excel
    xlsx_path = os.path.join(dashboard_dir, f'HOAi_Daily_Report_{report_date}.xlsx')
    print("\n[1/2] Generating Excel workbook...")
    generate_excel(data, xlsx_path)

    # Backward-compat flat copy
    flat_xlsx = os.path.join(OUTPUT_DIR, f'HOAi_Daily_Report_{report_date}.xlsx')
    import shutil
    shutil.copy2(xlsx_path, flat_xlsx)

    # Generate PDF
    pdf_path = os.path.join(dashboard_dir, f'HOAi_Daily_Report_{report_date}.pdf')
    print("[2/2] Generating PDF report...")
    generate_pdf(data, pdf_path)

    # Backward-compat flat copy for HTML (pdf fallback)
    flat_html = os.path.join(OUTPUT_DIR, f'HOAi_Daily_Report_{report_date}.html')
    html_in_dashboard = pdf_path.replace('.pdf', '.html')
    if os.path.exists(html_in_dashboard):
        shutil.copy2(html_in_dashboard, flat_html)

    # Generate per-company outreach pages
    date_dir = os.path.join(OUTPUT_DIR, month_dir, report_date)
    print("[3/4] Generating outreach pages...")
    outreach_count = generate_outreach_pages(data, date_dir)
    print(f"  Outreach: {outreach_count} company pages")

    # Generate overage notification emails
    print("[4/4] Generating overage emails...")
    overage_count = generate_overage_emails(data, date_dir)
    print(f"  Overage emails: {overage_count} companies")

    print(f"\nDone. Output in {dashboard_dir}/")


if __name__ == '__main__':
    main()
