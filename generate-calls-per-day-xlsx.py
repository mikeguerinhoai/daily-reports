"""Generate Calls/Day Excel workbook with Q4 2025 + Q1 2026 + April MTD.

Two tables on one sheet:
  1. Total Calls (raw values)
  2. Calls/Day (Excel formulas dividing table 1 by days-in-month)
"""
import json, sys

sys.stdout.reconfigure(encoding='utf-8')

with open('daily-reports/data/daily-report-2026-04-13.json') as f:
    d = json.load(f)

from supabase.db import get_cursor

with get_cursor() as cur:
    cur.execute('''
        SELECT
            mc.name as company,
            DATE_TRUNC('month', cl.start_time)::date as month,
            COUNT(*) as total_calls
        FROM call_logs cl
        JOIN management_company mc ON mc.id = cl.management_company_id
        WHERE cl.start_time >= '2025-10-01'
          AND cl.start_time < '2026-04-01'
          AND mc.deleted_at IS NULL
          AND (cl.channel IS NULL OR cl.channel != 'sms')
        GROUP BY mc.name, DATE_TRUNC('month', cl.start_time)::date
        ORDER BY mc.name, month
    ''')
    monthly_rows = cur.fetchall()

monthly = {}
for r in monthly_rows:
    co = r['company']
    m = str(r['month'])[:7]
    if co not in monthly:
        monthly[co] = {}
    monthly[co][m] = r['total_calls']

months = ['2025-10', '2025-11', '2025-12', '2026-01', '2026-02', '2026-03']
month_days_map = {'2025-10': 31, '2025-11': 30, '2025-12': 31, '2026-01': 31, '2026-02': 28, '2026-03': 31}
apr_days = d['platform']['voice_summary']['days_elapsed']
labels = ['Oct', 'Nov', 'Dec', 'Jan', 'Feb', 'Mar', 'Apr']
days_list = [31, 30, 31, 31, 28, 31, apr_days]  # days per column

companies = set(d.get('voice', {}).keys())
for co in monthly:
    companies.add(co)

# Build results sorted by Apr calls/day desc
results = []
for co in companies:
    apr_data = d.get('voice', {}).get(co, {})
    apr_calls = apr_data.get('total_calls', 0)
    co_monthly = monthly.get(co, {})
    vals = [co_monthly.get(m, 0) for m in months]
    if sum(vals) + apr_calls == 0:
        continue
    apr_d = round(apr_calls / apr_days, 1) if apr_days else 0
    results.append({'company': co, 'vals': vals, 'apr': apr_calls, 'apr_d': apr_d})

results.sort(key=lambda x: x['apr_d'], reverse=True)

# --- Excel ---
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'Voice Usage'

# Styles
header_font = Font(bold=True, color='FFFFFF', size=11)
header_fill = PatternFill(start_color='2D3748', end_color='2D3748', fill_type='solid')
q4_fill = PatternFill(start_color='EBF5FB', end_color='EBF5FB', fill_type='solid')
q1_fill = PatternFill(start_color='E8F8F5', end_color='E8F8F5', fill_type='solid')
apr_fill = PatternFill(start_color='FEF9E7', end_color='FEF9E7', fill_type='solid')
total_font = Font(bold=True, size=11)
total_fill = PatternFill(start_color='F0F0F0', end_color='F0F0F0', fill_type='solid')
active_fill = PatternFill(start_color='D4EFDF', end_color='D4EFDF', fill_type='solid')
thin_border = Border(bottom=Side(style='thin', color='D5D8DC'))
thick_border = Border(top=Side(style='medium', color='2D3748'), bottom=Side(style='medium', color='2D3748'))
section_font = Font(bold=True, size=13, color='1A5276')


def fill_for_col(col_idx):
    """Return fill based on column index (0-based): 0-2 = Q4, 3-5 = Q1, 6 = Apr."""
    if col_idx < 3:
        return q4_fill
    elif col_idx < 6:
        return q1_fill
    return apr_fill


def write_quarter_labels(ws, row):
    ws.cell(row=row, column=2, value='Q4 2025')
    ws.cell(row=row, column=2).font = Font(bold=True, size=11)
    ws.cell(row=row, column=2).alignment = Alignment(horizontal='center')
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
    ws.cell(row=row, column=5, value='Q1 2026')
    ws.cell(row=row, column=5).font = Font(bold=True, size=11)
    ws.cell(row=row, column=5).alignment = Alignment(horizontal='center')
    ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=7)
    ws.cell(row=row, column=8, value='Apr MTD')
    ws.cell(row=row, column=8).font = Font(bold=True, size=11)
    ws.cell(row=row, column=8).alignment = Alignment(horizontal='center')


def write_col_headers(ws, row):
    headers = ['Company'] + labels
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center' if c > 1 else 'left')


# ============================================================
# TABLE 1: Total Calls
# ============================================================
row = 1
ws.cell(row=row, column=1, value='Total Calls by Company')
ws.cell(row=row, column=1).font = section_font
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
row += 1

write_quarter_labels(ws, row)
row += 1

write_col_headers(ws, row)
header_row_t1 = row
row += 1

# Data rows — raw call counts
data_start_t1 = row
for r in results:
    ws.cell(row=row, column=1, value=r['company'])
    all_vals = r['vals'] + [r['apr']]
    for i, v in enumerate(all_vals):
        cell = ws.cell(row=row, column=i + 2, value=v if v > 0 else None)
        cell.number_format = '#,##0'
        cell.alignment = Alignment(horizontal='center')
        cell.fill = fill_for_col(i)
    for c in range(1, 9):
        ws.cell(row=row, column=c).border = thin_border
    row += 1
data_end_t1 = row - 1

# Book total (SUM formulas)
for c in range(1, 9):
    ws.cell(row=row, column=c).font = total_font
    ws.cell(row=row, column=c).fill = total_fill
    ws.cell(row=row, column=c).border = thick_border
ws.cell(row=row, column=1, value='BOOK TOTAL')
for col in range(2, 9):
    col_letter = get_column_letter(col)
    cell = ws.cell(row=row, column=col,
                   value=f'=SUM({col_letter}{data_start_t1}:{col_letter}{data_end_t1})')
    cell.number_format = '#,##0'
    cell.alignment = Alignment(horizontal='center')
total_row_t1 = row
row += 1

# Active customers (COUNTA formulas)
ws.cell(row=row, column=1, value='Active Customers')
ws.cell(row=row, column=1).font = Font(bold=True, color='1A5276')
for col in range(2, 9):
    col_letter = get_column_letter(col)
    cell = ws.cell(row=row, column=col,
                   value=f'=COUNTA({col_letter}{data_start_t1}:{col_letter}{data_end_t1})')
    cell.font = Font(bold=True, size=12, color='1A5276')
    cell.alignment = Alignment(horizontal='center')
    cell.fill = active_fill
row += 1

# Days in month reference row
ws.cell(row=row, column=1, value='Days in Month')
ws.cell(row=row, column=1).font = Font(italic=True, color='999999')
days_row = row
for i, dys in enumerate(days_list):
    cell = ws.cell(row=row, column=i + 2, value=dys)
    cell.font = Font(italic=True, color='999999')
    cell.alignment = Alignment(horizontal='center')
row += 1

# ============================================================
# TABLE 2: Calls/Day (formulas referencing Table 1)
# ============================================================
row += 1  # blank spacer row

ws.cell(row=row, column=1, value='Calls / Day by Company')
ws.cell(row=row, column=1).font = section_font
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
row += 1

write_quarter_labels(ws, row)
row += 1

write_col_headers(ws, row)
row += 1

# Data rows — formulas: =IF(T1_cell="","",T1_cell/days)
data_start_t2 = row
for idx, r in enumerate(results):
    ws.cell(row=row, column=1, value=r['company'])
    src_row = data_start_t1 + idx  # corresponding row in table 1
    for col in range(2, 9):
        col_letter = get_column_letter(col)
        days_ref = f'${col_letter}${days_row}'
        src_ref = f'{col_letter}{src_row}'
        formula = f'=IF({src_ref}="","",{src_ref}/{days_ref})'
        cell = ws.cell(row=row, column=col, value=formula)
        cell.number_format = '#,##0.0'
        cell.alignment = Alignment(horizontal='center')
        cell.fill = fill_for_col(col - 2)
    for c in range(1, 9):
        ws.cell(row=row, column=c).border = thin_border
    row += 1
data_end_t2 = row - 1

# Book total (formula: total_calls / days)
for c in range(1, 9):
    ws.cell(row=row, column=c).font = total_font
    ws.cell(row=row, column=c).fill = total_fill
    ws.cell(row=row, column=c).border = thick_border
ws.cell(row=row, column=1, value='BOOK TOTAL')
for col in range(2, 9):
    col_letter = get_column_letter(col)
    days_ref = f'${col_letter}${days_row}'
    total_ref = f'{col_letter}{total_row_t1}'
    cell = ws.cell(row=row, column=col, value=f'={total_ref}/{days_ref}')
    cell.number_format = '#,##0.0'
    cell.alignment = Alignment(horizontal='center')
row += 1

# Active customers (same formula, reference table 1)
ws.cell(row=row, column=1, value='Active Customers')
ws.cell(row=row, column=1).font = Font(bold=True, color='1A5276')
for col in range(2, 9):
    col_letter = get_column_letter(col)
    cell = ws.cell(row=row, column=col,
                   value=f'=COUNTA({col_letter}{data_start_t1}:{col_letter}{data_end_t1})')
    cell.font = Font(bold=True, size=12, color='1A5276')
    cell.alignment = Alignment(horizontal='center')
    cell.fill = active_fill

# Column widths
ws.column_dimensions['A'].width = 42
for c in range(2, 9):
    ws.column_dimensions[get_column_letter(c)].width = 12

ws.freeze_panes = 'B4'

outpath = 'daily-reports/output/HOAi_Voice_Calls_Per_Day_Q4Q1_v2.xlsx'
wb.save(outpath)
print(f'Saved: {outpath}')
print()
print('Active Customers by Month:')
for i, lbl in enumerate(labels[:6]):
    count = sum(1 for r in results if r['vals'][i] > 0)
    print(f'  {lbl}: {count}')
apr_count = sum(1 for r in results if r['apr'] > 0)
print(f'  Apr: {apr_count}')
